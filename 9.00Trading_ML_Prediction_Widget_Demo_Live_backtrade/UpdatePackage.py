#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          TRADING APP — AUTOMATED UPDATE PACKAGE  v2.0                        ║
║ ─────────────────────────────────────────────────────────────────────────── ║
║  Place this file in your project root, then run:                             ║
║      python update_v2.py                                                     ║
║                                                                              ║
║  PATCHES APPLIED                                                             ║
║  App_MACD_AI_HybridScore_Latest1.py                                                                 ║
║    ADD  stop_trading()                                                       ║
║    ADD  predict_future_trend()                                               ║
║    ADD  close_partial()                                                      ║
║    ADD  Trading Time Window system  (14 methods)                             ║
║    MOD  start_trading()   — time-window aware replacement                    ║
║    MOD  trading_loop()    — time-window guard inserted                       ║
║    MOD  __init__()        — time-config lines appended                       ║
║                                                                              ║
║  scalping_strategy.py                                                        ║
║    ADD  ScalpingStrategy.execute_buy()                                       ║
║    ADD  ScalpingStrategy.execute_sell()                                      ║
║    ADD  ScalpingStrategy.get_strategy_info()                                 ║
║    ADD  BacktestScalpingStrategy.export_trades_to_excel()                    ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os
import re
import sys
import shutil
import textwrap
from pathlib import Path
from datetime import datetime

# ════════════════════════════════════════════════════════════════════════════
# CONSOLE HELPERS
# ════════════════════════════════════════════════════════════════════════════

GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

def ok(msg):   print(f"  {GREEN}✅ {msg}{RESET}")
def warn(msg): print(f"  {YELLOW}⚠️  {msg}{RESET}")
def err(msg):  print(f"  {RED}❌ {msg}{RESET}")
def info(msg): print(f"  {CYAN}ℹ️  {msg}{RESET}")
def hdr(msg):  print(f"\n{BOLD}{CYAN}{'─'*70}\n  {msg}\n{'─'*70}{RESET}")


# ════════════════════════════════════════════════════════════════════════════
# FILE PATCHER
# ════════════════════════════════════════════════════════════════════════════

class FilePatcher:
    """Safe, idempotent file patcher with backup and rollback."""

    def __init__(self, filepath: str):
        self.path     = Path(filepath)
        self.original = ""
        self.content  = ""
        self.log: list[str] = []

    # ── I/O ─────────────────────────────────────────────────────────────
    def load(self) -> bool:
        if not self.path.exists():
            err(f"File not found: {self.path}"); return False
        with open(self.path, "r", encoding="utf-8") as f:
            self.content = f.read()
        self.original = self.content
        ok(f"Loaded  {self.path.name}  ({len(self.content):,} chars)"); return True

    def backup(self) -> Path:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = self.path.with_suffix(f".backup_{ts}.py")
        shutil.copy2(self.path, dest)
        ok(f"Backup → {dest.name}"); return dest

    def save(self):
        with open(self.path, "w", encoding="utf-8") as f:
            f.write(self.content)
        ok(f"Saved  {self.path.name}")

    def rollback(self):
        with open(self.path, "w", encoding="utf-8") as f:
            f.write(self.original)
        warn(f"Rolled back {self.path.name}")

    # ── Query helpers ────────────────────────────────────────────────────
    def has_method(self, name: str) -> bool:
        return bool(re.search(rf"^\s+def {re.escape(name)}\s*\(", self.content, re.M))

    def has_anchor(self, anchor: str) -> bool:
        return anchor in self.content

    # ── Patching primitives ──────────────────────────────────────────────
    def insert_after(self, anchor: str, code: str, description: str = "") -> bool:
        """Insert *code* immediately after the first occurrence of *anchor*."""
        if anchor not in self.content:
            err(f"Anchor not found for: {description or anchor[:60]}"); return False
        pos = self.content.find(anchor) + len(anchor)
        self.content = self.content[:pos] + "\n" + code + self.content[pos:]
        self.log.append(description or "insert_after")
        ok(description or "insert_after"); return True

    def insert_before(self, anchor: str, code: str, description: str = "") -> bool:
        """Insert *code* immediately before the first occurrence of *anchor*."""
        if anchor not in self.content:
            err(f"Anchor not found for: {description or anchor[:60]}"); return False
        pos = self.content.find(anchor)
        self.content = self.content[:pos] + code + "\n" + self.content[pos:]
        self.log.append(description or "insert_before")
        ok(description or "insert_before"); return True

    def replace_block(self, old: str, new: str, description: str = "") -> bool:
        """Replace an exact block of text."""
        if old not in self.content:
            err(f"Block not found for: {description or old[:60]}"); return False
        self.content = self.content.replace(old, new, 1)
        self.log.append(description or "replace_block")
        ok(description or "replace_block"); return True

    def find_method_span(self, method_name: str, class_indent: int = 4) -> tuple:
        """
        Return (start, end) character indices of a method definition inside a class.
        The span covers from 'def method_name' through the last line before the
        next same-level def/decorator/class, or EOF.
        Returns (-1, -1) if not found.
        """
        ind   = " " * class_indent
        pat   = re.compile(rf"^{re.escape(ind)}def {re.escape(method_name)}\s*\(", re.M)
        m     = pat.search(self.content)
        if not m:
            return -1, -1

        start = m.start()
        # Walk lines after start; stop when we hit a new def/decorator/class at
        # the same or lower indentation
        rest  = self.content[start:]
        lines = rest.split("\n")
        end_offset = len(rest)       # default: to EOF

        for i, line in enumerate(lines):
            if i == 0:
                continue
            if not line.strip():     # blank line — keep going
                continue
            cur_indent = len(line) - len(line.lstrip())
            if cur_indent <= class_indent and line.lstrip()[:1] in ("d", "@", "c", "#"):
                # Next def / decorator / class at same level
                end_offset = len("\n".join(lines[:i]))
                break

        return start, start + end_offset

    def replace_method(self, method_name: str, new_code: str,
                       class_indent: int = 4, description: str = "") -> bool:
        """Replace an entire method definition."""
        s, e = self.find_method_span(method_name, class_indent)
        if s == -1:
            err(f"Method not found: {method_name}"); return False
        self.content = self.content[:s] + new_code.rstrip() + "\n\n" + self.content[e:]
        self.log.append(description or f"replace_method({method_name})")
        ok(description or f"replace_method({method_name})"); return True


# ════════════════════════════════════════════════════════════════════════════
# ── CODE BLOCKS:  App_MACD_AI_HybridScore_Latest1.py
# ════════════════════════════════════════════════════════════════════════════

# ── 1. stop_trading ──────────────────────────────────────────────────────────
_STOP_TRADING = '''\
    def stop_trading(self):
        """Stop the trading loop and clean up state."""
        if not self.running:
            self.log_message("⚠️ Trading is not currently running.", "orange")
            return

        self.running         = False
        self.trading_running = False

        # Cancel any scheduled time-window jobs
        if hasattr(self, '_cancel_scheduled_trading'):
            self._cancel_scheduled_trading()

        if hasattr(self, 'timer') and self.timer:
            try:
                self.timer.stop()
            except Exception:
                pass

        if hasattr(self, 'start_btn') and self.start_btn:
            self.start_btn.config(state=tk.NORMAL)
        if hasattr(self, 'stop_btn') and self.stop_btn:
            self.stop_btn.config(state=tk.DISABLED)

        self.update_mode_display(self.mode_var.get())
        self.update_status_indicators("parking")
        self.log_message("🛑 Trading stopped.", "orange")

        try:
            import pyttsx3
            engine = pyttsx3.init()
            engine.setProperty('rate', 135)
            engine.say("Trading stopped")
            engine.runAndWait()
        except Exception:
            pass

        self.play_notification("tick")
        self.update_stats()

'''

# ── 2. predict_future_trend ──────────────────────────────────────────────────
_PREDICT_FUTURE = '''\
    def predict_future_trend(self, n_future: int = 5):
        """
        Use the current ML model to predict the trend for the next n_future candles.
        Returns (predictions: list, confidence: float).
        """
        if not self.ml_enabled or self.current_ml_model is None:
            return [], 0.0

        if not getattr(self.current_ml_model, 'is_trained', False):
            return [], 0.0

        df = self.get_market_data()
        if df is None or df.empty:
            return [], 0.0

        try:
            if hasattr(self.strategy, 'calculate_indicators'):
                df = self.strategy.calculate_indicators(df)

            conf, prediction, _ = self.current_ml_model.predict(df, n_future)
            confidence = float(conf) if conf <= 1.0 else float(conf) / 100.0

            if prediction == 1:
                predictions = ['bullish'] * n_future
            elif prediction == -1:
                predictions = ['bearish'] * n_future
            else:
                predictions = ['neutral'] * n_future

            return predictions, confidence

        except Exception as e:
            self.log_message(f"⚠️ predict_future_trend error: {e}", "orange")
            return [], 0.0

'''

# ── 3. close_partial ─────────────────────────────────────────────────────────
_CLOSE_PARTIAL = '''\
    def close_partial(self, fraction: float = 0.5):
        """
        Close a fraction of the current position.
        fraction : 0 < fraction <= 1
        """
        if self.position.get('type') is None or self.position.get('quantity') is None:
            self.log_message("⚠️ close_partial: no open position.", "orange")
            return False

        if not (0 < fraction <= 1):
            self.log_message(f"❌ close_partial: invalid fraction {fraction}", "red")
            return False

        qty_to_close = self.position['quantity'] * fraction
        current_price = self.get_current_price()
        if current_price is None:
            if self.current_data is not None:
                current_price = float(self.current_data.get('Close', 0))
        if not current_price:
            self.log_message("❌ close_partial: cannot determine price.", "red")
            return False

        self.log_message(
            f"📤 Partial close: {fraction*100:.0f}% "
            f"({qty_to_close:.4f} units) @ ${current_price:.4f}", "blue")

        success = self.place_order(
            'sell', current_price, quantity=qty_to_close,
            exit_reason='partial_profit_target')

        if success:
            remaining = self.position['quantity'] - qty_to_close
            if remaining > 1e-8:
                self.position['quantity'] = remaining
                self.log_message(
                    f"✅ Partial close complete. Remaining: {remaining:.4f}", "green")
            else:
                self.position = {
                    'type': None, 'price': None, 'quantity': None, 'time': None,
                    'stop_loss': None, 'trailing_stop': None, 'entry_confidence': None
                }
                self.update_status_indicators("parking")
                self.log_message("✅ Position fully closed via close_partial.", "green")

        return success

'''

# ── 4. Trading Time Window — 14 methods ──────────────────────────────────────
_TIME_WINDOW_METHODS = '''\
    # ═══════════════════════════════════════════════════════════════════════
    # TRADING TIME WINDOW
    # ═══════════════════════════════════════════════════════════════════════

    def _init_trading_time_config(self) -> dict:
        """Per-strategy time window.  start==end → no restriction."""
        cfg = {}
        for strategy in ('Momentum', 'Kalman', 'Scalping'):
            cfg[strategy] = {
                'start_h': tk.IntVar(value=0),
                'start_m': tk.IntVar(value=0),
                'end_h':   tk.IntVar(value=0),
                'end_m':   tk.IntVar(value=0),
            }
        return cfg

    def _is_time_unconstrained(self, strategy: str) -> bool:
        cfg = self.trading_time_config.get(strategy, {})
        return (cfg['start_h'].get() == cfg['end_h'].get() and
                cfg['start_m'].get() == cfg['end_m'].get())

    def _get_window_minutes(self, strategy: str) -> tuple:
        cfg = self.trading_time_config[strategy]
        s = cfg['start_h'].get() * 60 + cfg['start_m'].get()
        e = cfg['end_h'].get()   * 60 + cfg['end_m'].get()
        return s, e

    def _is_within_trading_window(self, strategy: str) -> bool:
        if self._is_time_unconstrained(strategy):
            return True
        now_utc  = datetime.now(timezone.utc)
        now_min  = now_utc.hour * 60 + now_utc.minute
        s, e     = self._get_window_minutes(strategy)
        if s < e:
            return s <= now_min < e
        return now_min >= s or now_min < e   # overnight window

    def _seconds_until_start(self, strategy: str) -> float:
        now_utc  = datetime.now(timezone.utc)
        now_min  = now_utc.hour * 60 + now_utc.minute
        now_sec  = now_utc.second
        s, _     = self._get_window_minutes(strategy)
        diff_min = (s - now_min) % (24 * 60)
        return max(0., diff_min * 60 - now_sec)

    def _seconds_until_end(self, strategy: str) -> float:
        now_utc  = datetime.now(timezone.utc)
        now_min  = now_utc.hour * 60 + now_utc.minute
        now_sec  = now_utc.second
        _, e     = self._get_window_minutes(strategy)
        diff_min = (e - now_min) % (24 * 60)
        return max(0., diff_min * 60 - now_sec)

    def _cancel_scheduled_trading(self):
        for attr in ('_sched_start_id', '_sched_stop_id'):
            aid = getattr(self, attr, None)
            if aid is not None:
                try:
                    self.root.after_cancel(aid)
                except Exception:
                    pass
                setattr(self, attr, None)
        self._waiting_to_start = False

    def _do_start_trading(self):
        """Spin up the trading thread (internal — called by start_trading)."""
        self._waiting_to_start = False
        if hasattr(self, 'trading_thread') and self.trading_thread is not None:
            if self.trading_thread.is_alive():
                self.running = False
                self.trading_running = False
                self.trading_thread.join(timeout=3.0)
            self.trading_thread = None

        self.running         = True
        self.trading_running = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.update_mode_display(self.mode_var.get())
        self.trading_thread = threading.Thread(
            target=self.trading_loop, daemon=True, name="TradingLoop")
        self.trading_thread.start()
        self.log_message("▶ Trading started.", "green")
        self.play_notification("tick")
        self.enable_ai_button()
        try:
            import pyttsx3
            e = pyttsx3.init(); e.setProperty('rate', 135)
            e.say("Trading started"); e.runAndWait()
        except Exception:
            pass

    def _scheduled_start_callback(self):
        self._sched_start_id = None
        if not self._waiting_to_start:
            return
        strategy = self.strategy_type_var.get()
        now_str  = datetime.now(timezone.utc).strftime('%H:%M UTC')
        self.log_message(
            f"⏰ [{strategy}] Scheduled start at {now_str}", "green")
        self._do_start_trading()
        if not self._is_time_unconstrained(strategy):
            stop_secs = self._seconds_until_end(strategy)
            if stop_secs > 0:
                self._sched_stop_id = self.root.after(
                    int(stop_secs * 1000), self._scheduled_stop_callback)
                _, em = self._get_window_minutes(strategy)
                eh, em2 = divmod(em, 60)
                self.log_message(
                    f"⏰ [{strategy}] Auto-stop in "
                    f"{stop_secs/60:.1f} min ({eh:02d}:{em2:02d} UTC)", "blue")

    def _scheduled_stop_callback(self):
        self._sched_stop_id = None
        if not self.running:
            return
        strategy = self.strategy_type_var.get()
        now_str  = datetime.now(timezone.utc).strftime('%H:%M UTC')
        self.log_message(
            f"⏰ [{strategy}] Trading window closed at {now_str}.", "orange")
        self.stop_trading()

    def _update_wait_countdown(self, strategy: str, sh: int, sm: int):
        if not getattr(self, '_waiting_to_start', False):
            return
        secs = self._seconds_until_start(strategy)
        if secs <= 0:
            return
        mins = int(secs / 60)
        self.mode_display.config(
            text=f"⏰ STARTS IN {mins}m → {sh:02d}:{sm:02d} UTC",
            foreground='orange')
        self.root.after(60_000,
            lambda: self._update_wait_countdown(strategy, sh, sm))

    def open_time_settings_panel(self, strategy: str):
        """Open a floating panel to set trading hours for *strategy*."""
        attr     = f"_time_panel_{strategy}"
        existing = getattr(self, attr, None)
        if existing and existing.winfo_exists():
            existing.lift(); existing.focus_force(); return

        cfg   = self.trading_time_config[strategy]
        panel = tk.Toplevel(self.root)
        panel.title(f"⏰ Trading Hours — {strategy}")
        panel.geometry("390x285")
        panel.resizable(False, False)
        panel.grab_set()
        setattr(self, attr, panel)

        hdr_f = tk.Frame(panel, bg="#1a1a2e", pady=8)
        hdr_f.pack(fill=tk.X)
        tk.Label(hdr_f, text=f"⏰  {strategy} — Trading Hours (UTC)",
                 bg="#1a1a2e", fg="white",
                 font=("Arial", 12, "bold")).pack()

        body = ttk.Frame(panel, padding=15)
        body.pack(fill=tk.BOTH, expand=True)

        def _time_row(lbl_text, h_var, m_var, row):
            ttk.Label(body, text=lbl_text,
                      font=("Arial", 10, "bold")).grid(
                row=row, column=0, sticky="w", pady=6, padx=5)
            sb_h = ttk.Spinbox(body, from_=0, to=23, textvariable=h_var,
                               width=5, format="%02.0f", font=("Consolas", 11))
            sb_h.grid(row=row, column=1, padx=(10, 2))
            ttk.Label(body, text=":", font=("Consolas", 13, "bold")).grid(
                row=row, column=2)
            sb_m = ttk.Spinbox(body, from_=0, to=59, textvariable=m_var,
                               width=5, format="%02.0f", font=("Consolas", 11))
            sb_m.grid(row=row, column=3, padx=(2, 10))
            ttk.Label(body, text="UTC", foreground="grey").grid(
                row=row, column=4, padx=5)

        _time_row("▶  Start time :", cfg["start_h"], cfg["start_m"], row=0)
        _time_row("⏹  End   time :", cfg["end_h"],   cfg["end_m"],   row=1)

        status_var = tk.StringVar()
        status_lbl = ttk.Label(body, textvariable=status_var,
                               font=("Arial", 9))
        status_lbl.grid(row=2, column=0, columnspan=5,
                        pady=(10, 0), sticky="w")

        def _refresh(*_):
            sh, sm = cfg["start_h"].get(), cfg["start_m"].get()
            eh, em = cfg["end_h"].get(),   cfg["end_m"].get()
            if sh == eh and sm == em:
                status_var.set("ℹ️  No restriction — trades run at any time")
                status_lbl.config(foreground="#0066CC")
            else:
                status_var.set(
                    f"✅  Window: {sh:02d}:{sm:02d} → {eh:02d}:{em:02d} UTC")
                status_lbl.config(foreground="#006600")

        for v in (cfg["start_h"], cfg["start_m"], cfg["end_h"], cfg["end_m"]):
            v.trace_add("write", _refresh)
        _refresh()

        btn_f = ttk.Frame(panel, padding=(15, 5, 15, 15))
        btn_f.pack(fill=tk.X)

        def _reset():
            cfg["end_h"].set(cfg["start_h"].get())
            cfg["end_m"].set(cfg["start_m"].get())
            self.log_message(
                f"⏰ [{strategy}] Trading hours reset — no constraint.", "blue")

        def _save():
            _refresh()
            sh, sm = cfg["start_h"].get(), cfg["start_m"].get()
            eh, em = cfg["end_h"].get(),   cfg["end_m"].get()
            if sh == eh and sm == em:
                self.log_message(
                    f"⏰ [{strategy}] Trading hours: unrestricted", "blue")
            else:
                self.log_message(
                    f"⏰ [{strategy}] Trading hours: "
                    f"{sh:02d}:{sm:02d} → {eh:02d}:{em:02d} UTC", "green")
            panel.destroy()

        tk.Button(btn_f, text="🔄 Reset (no limit)", command=_reset,
                  bg="#FFA500", fg="white", font=("Arial", 9, "bold"),
                  relief="raised", bd=2, cursor="hand2",
                  padx=10, pady=4).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_f, text="💾 Save & Close", command=_save,
                  bg="#0066CC", fg="white", font=("Arial", 9, "bold"),
                  relief="raised", bd=2, cursor="hand2",
                  padx=10, pady=4).pack(side=tk.LEFT, padx=4)
        tk.Button(btn_f, text="❌ Cancel", command=panel.destroy,
                  bg="#CC0000", fg="white", font=("Arial", 9, "bold"),
                  relief="raised", bd=2, cursor="hand2",
                  padx=10, pady=4).pack(side=tk.RIGHT, padx=4)

    def _add_time_settings_button(self, parent_frame, strategy: str):
        """Inject a ⏰ Trading Hours row at the top of a settings tab."""
        cfg = self.trading_time_config.get(strategy)
        if cfg is None:
            return

        row = tk.Frame(parent_frame, bg="#f0f4ff", pady=4)
        row.pack(fill=tk.X, padx=5, pady=(4, 0))

        def _summary():
            sh, sm = cfg["start_h"].get(), cfg["start_m"].get()
            eh, em = cfg["end_h"].get(),   cfg["end_m"].get()
            if sh == eh and sm == em:
                return "⏰ Trading hours: unrestricted"
            return f"⏰ Trading hours: {sh:02d}:{sm:02d} → {eh:02d}:{em:02d} UTC"

        lbl_var = tk.StringVar(value=_summary())
        lbl = tk.Label(row, textvariable=lbl_var, bg="#f0f4ff",
                       font=("Arial", 9), anchor="w")
        lbl.pack(side=tk.LEFT, padx=8, fill=tk.X, expand=True)

        def _refresh(*_):
            lbl_var.set(_summary())
            sh, sm = cfg["start_h"].get(), cfg["start_m"].get()
            eh, em = cfg["end_h"].get(),   cfg["end_m"].get()
            lbl.config(fg="#006600"
                       if not (sh == eh and sm == em) else "#444444")

        for v in (cfg["start_h"], cfg["start_m"], cfg["end_h"], cfg["end_m"]):
            v.trace_add("write", _refresh)

        tk.Button(row, text="⚙ Set Hours",
                  command=lambda s=strategy: self.open_time_settings_panel(s),
                  bg="#0066CC", fg="white", font=("Arial", 8, "bold"),
                  relief="raised", bd=2, cursor="hand2",
                  padx=8, pady=2).pack(side=tk.RIGHT, padx=8)

'''

# ── 5. start_trading replacement ─────────────────────────────────────────────
_START_TRADING = '''\
    def start_trading(self):
        """Start trading, respecting the per-strategy time window."""
        self.update_status_indicators("parking")
        self.order_size_pct    = self.order_size_var.get()
        self.stop_loss_pct     = self.stop_loss_var.get() / 100
        self.trailing_stop_pct = self.trailing_stop_var.get() / 100
        self.trade_direction   = self.trade_direction_var.get()

        mode     = self.mode_var.get().lower()
        strategy = self.strategy_type_var.get()

        if mode != "backtest":
            if not hasattr(self, 'market_api') or self.market_api is None:
                messagebox.showerror("Error", "Please check connection first!")
                return

        if self.running:
            return

        # Cancel any previously scheduled jobs
        if hasattr(self, '_cancel_scheduled_trading'):
            self._cancel_scheduled_trading()

        # ── Time-window logic (live / demo only) ────────────────────────
        if (mode in ('live', 'demo')
                and hasattr(self, 'trading_time_config')
                and not self._is_time_unconstrained(strategy)):
            s, e   = self._get_window_minutes(strategy)
            sh, sm = divmod(s, 60)
            eh, em = divmod(e, 60)

            if self._is_within_trading_window(strategy):
                self.log_message(
                    f"⏰ [{strategy}] Inside window "
                    f"({sh:02d}:{sm:02d}→{eh:02d}:{em:02d} UTC) — starting.", "green")
                self._do_start_trading()
                stop_secs = self._seconds_until_end(strategy)
                if stop_secs > 0:
                    self._sched_stop_id = self.root.after(
                        int(stop_secs * 1000), self._scheduled_stop_callback)
                    self.log_message(
                        f"⏰ [{strategy}] Auto-stop in "
                        f"{stop_secs/60:.1f} min at {eh:02d}:{em:02d} UTC", "blue")
            else:
                wait_secs = self._seconds_until_start(strategy)
                self._waiting_to_start = True
                self.log_message(
                    f"⏰ [{strategy}] Outside window. "
                    f"Waiting {wait_secs/60:.1f} min until "
                    f"{sh:02d}:{sm:02d} UTC …", "orange")
                self.mode_display.config(
                    text=f"⏰ WAITING {sh:02d}:{sm:02d} UTC",
                    foreground="orange")
                self._sched_start_id = self.root.after(
                    int(wait_secs * 1000), self._scheduled_start_callback)
                self.stop_btn.config(state=tk.NORMAL)
                self.start_btn.config(state=tk.DISABLED)
                self._update_wait_countdown(strategy, sh, sm)
            return

        # ── No time constraint → start immediately ───────────────────────
        self._do_start_trading()
        self.log_message(
            f"📊 {strategy} — Order {self.order_size_pct}% | "
            f"SL {self.stop_loss_pct*100:.1f}% | "
            f"Trail {self.trailing_stop_pct*100:.1f}% | "
            f"Dir {self.trade_direction.upper()}", "blue")

'''

# ── 6. __init__ time-config additions (inserted after weight_manager line) ───
_INIT_TIME_CONFIG = """
        # ── Trading Time Window ──────────────────────────────────────────────
        self.trading_time_config  = self._init_trading_time_config()
        self._sched_start_id      = None
        self._sched_stop_id       = None
        self._waiting_to_start    = False
"""

# ── 7. trading_loop time-window guard ─────────────────────────────────────────
#   Inserted immediately after `df = self.get_market_data()` in trading_loop
_TRADING_LOOP_PATCH = """
                # ── Time-window guard ────────────────────────────────────────
                _strategy_name = self.strategy_type_var.get()
                _mode_name     = self.mode_var.get().lower()
                if (_mode_name in ('live', 'demo')
                        and hasattr(self, 'trading_time_config')
                        and not self._is_time_unconstrained(_strategy_name)
                        and not self._is_within_trading_window(_strategy_name)):
                    self.log_message(
                        f"⏰ [{_strategy_name}] Outside trading window — pausing.",
                        "orange")
                    _wait = min(self._seconds_until_start(_strategy_name), 300)
                    time.sleep(max(5, _wait))
                    continue
                # ─────────────────────────────────────────────────────────────
"""

# ── 8. create_momentum_parameter_controls hook ────────────────────────────────
_MOMENTUM_TIME_HOOK = '        self._add_time_settings_button(parent, "Momentum")\n'
_KALMAN_TIME_HOOK   = '        self._add_time_settings_button(parent, "Kalman")\n'
_SCALPING_TIME_HOOK = '        self._add_time_settings_button(parent, "Scalping")\n'


# ════════════════════════════════════════════════════════════════════════════
# ── CODE BLOCKS:  scalping_strategy.py
# ════════════════════════════════════════════════════════════════════════════

_SCALPING_EXECUTE_BUY = '''\
    def execute_buy(self, shares, price, atr=1., quality_score=70, tier=1, **kwargs):
        """Open long/short from pending signal. Returns (success, qty, order_id)."""
        direction = (self._pending_signal.get('direction', 'long')
                     if self._pending_signal else 'long')

        side    = 'buy' if direction == 'long' else 'sell'
        success = self.trading_app.place_order(
            side, price, quantity=shares, atr=atr, confidence=quality_score)

        if success:
            atr_mult = getattr(self, 'stop_loss_atr_mult', 2.2)
            stop = (price - atr * atr_mult if direction == 'long'
                    else price + atr * atr_mult)
            self.position = {
                'type': direction, 'entry_price': price,
                'quantity': shares, 'original_quantity': shares,
                'stop_loss': stop, 'trailing_stop': stop,
                'trailing_activated': False,
                'highest_price': price if direction == 'long' else None,
                'lowest_price':  price if direction == 'short' else None,
                'entry_bar': self.bar_count, 'partial_exits': 0,
                'tier': tier, 'entry_time': datetime.now(timezone.utc),
                'entry_quality_score': quality_score, 'entry_reason': 'signal',
                'trade_id': self.trade_counter + 1, 'partial_pnl_realised': 0.,
            }
            self.trade_counter += 1
            self.bars_held = 0
            self._transition_to_in_trade()
            self._log(
                f"✅ {'LONG' if direction=='long' else 'SHORT'} OPENED "
                f"T{tier} Q={quality_score} @${price:.4f} SL=${stop:.4f}",
                "green" if direction == 'long' else "red")
            return True, shares, self.trade_counter
        return False, 0, None

'''

_SCALPING_EXECUTE_SELL = '''\
    def execute_sell(self, reason="manual", exit_percentage=1.0, **kwargs):
        """Close all or part of the current position. Returns (success, profit, price)."""
        from datetime import datetime, timezone
        if self.strategy_state != StrategyState.IN_TRADE:
            self._log("⚠️ execute_sell: no open position", "orange")
            return False, 0., 0.

        current_price = self.trading_app.get_current_price()
        if current_price is None and self.trading_app.current_data is not None:
            current_price = float(self.trading_app.current_data.get('Close', 0))
        if not current_price:
            self._log("❌ execute_sell: cannot determine price", "red")
            return False, 0., 0.

        qty       = self.position.get('quantity', 0)
        close_qty = qty * exit_percentage
        ep        = self.position.get('entry_price', current_price)
        ptype     = self.position.get('type', 'long')

        side    = 'sell' if ptype == 'long' else 'buy'
        success = self.trading_app.place_order(
            side, current_price, quantity=close_qty, exit_reason=reason)

        if success:
            profit = ((current_price - ep) * close_qty if ptype == 'long'
                      else (ep - current_price) * close_qty)

            if exit_percentage >= 1.0:
                self.record_trade(
                    profit=profit + self.position.get('partial_pnl_realised', 0.),
                    exit_reason=reason, tier=self.position.get('tier'),
                    size=qty, direction=ptype,
                    entry_quality=self.position.get('entry_quality_score'),
                    entry_price=ep, exit_price=current_price,
                    hold_duration=self.bars_held,
                    entry_bar=self.position.get('entry_bar'),
                    exit_bar=self.bar_count)
                pnl_pct = profit / (ep * qty) * 100 if ep and qty else 0.
                self._log(
                    f"{'✅' if profit>0 else '❌'} {'LONG' if ptype=='long' else 'SHORT'} "
                    f"CLOSED @${current_price:.4f} P&L ${profit:+.2f} "
                    f"({pnl_pct:+.2f}%) — {reason}",
                    "green" if profit > 0 else "red")
                self.position = {
                    'type': None, 'entry_price': None, 'quantity': None,
                    'stop_loss': None, 'trailing_stop': None,
                    'trailing_activated': False, 'highest_price': None,
                    'lowest_price': None, 'entry_bar': None,
                    'partial_exits': 0, 'original_quantity': None, 'tier': None,
                    'entry_time': None, 'entry_quality_score': None,
                    'entry_reason': None, 'trade_id': None,
                    'partial_pnl_realised': 0.,
                }
                self.bars_held = 0
                self._transition_to_seeking_entry()
                self.trading_app.update_status_indicators("parking")
            else:
                remaining = qty - close_qty
                self.position['quantity']            = remaining
                self.position['partial_exits']       = self.position.get('partial_exits', 0) + 1
                self.position['partial_pnl_realised']= (
                    self.position.get('partial_pnl_realised', 0.) + profit)
                self._log(
                    f"📊 PARTIAL {exit_percentage:.0%} @${current_price:.4f} "
                    f"P&L ${profit:+.2f} | remaining {remaining:.4f}", "cyan")

            return True, profit, current_price
        return False, 0., current_price

'''

_SCALPING_GET_STRATEGY_INFO = '''\
    def get_strategy_info(self) -> dict:
        """Return metadata displayed in the strategy switch panel."""
        return {
            'name':                   "Professional Scalping v1.6.1",
            'version':                "1.6.1",
            'tier_system':            f"Tier 1 (Q≥{getattr(self,'quality_tier1_min',72)}) "
                                      f"/ Tier 2 (Q≥{getattr(self,'quality_min_long',55)})",
            'expected_trades_monthly':"30-50",
            'target_win_rate':        "47-55%",
            'target_cagr':            "35-60%",
            'target_sharpe':          "1.2-2.0",
            'max_drawdown':           "<10%",
            'tier1_description':      f"High-conviction Q≥{getattr(self,'quality_tier1_min',72)} "
                                      f"full EMA stack + all filters",
            'tier2_description':      f"Standard Q≥{getattr(self,'quality_min_long',55)} "
                                      f"entry — partial size",
            'timeframe':              getattr(self, 'timeframe', '1h'),
            'trade_direction':        getattr(self, 'trade_direction', 'both'),
        }

'''

_BACKTEST_EXPORT_TRADES = '''\
    def export_trades_to_excel(self, filename: str) -> bool:
        """Export trade_history to Excel with LONG/SHORT column."""
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill

            if not self.trade_history:
                print("⚠️ export_trades_to_excel: no trades"); return False

            rows = []
            for i, t in enumerate(self.trade_history, start=1):
                profit    = t.get('profit', 0.)
                ep        = t.get('entry_price', 0.)
                xp        = t.get('exit_price', 0.)
                sz        = t.get('size', 0.)
                direction = t.get('direction', 'long')
                entry_val = ep * sz if ep and sz else 0.
                ret_pct   = profit / entry_val * 100 if entry_val else 0.
                ts        = t.get('timestamp', '')
                if hasattr(ts, 'strftime'):
                    ts = ts.strftime('%Y-%m-%d %H:%M:%S')
                rows.append({
                    'Trade_#':       i,
                    'Direction':     direction.upper(),
                    'Entry_Bar':     t.get('entry_bar', ''),
                    'Exit_Bar':      t.get('exit_bar', ''),
                    'Timestamp':     ts,
                    'Entry_Price':   round(ep, 4),
                    'Exit_Price':    round(xp, 4),
                    'Size':          round(sz, 4),
                    'Profit_$':      round(profit, 2),
                    'Return_%':      round(ret_pct, 2),
                    'Hold_Bars':     t.get('hold_duration', 0),
                    'Exit_Reason':   t.get('exit_reason', ''),
                    'Tier':          t.get('tier', ''),
                    'Quality_Score': t.get('entry_quality', ''),
                    'Win':           'Yes' if profit > 0 else 'No',
                })

            df_trades = pd.DataFrame(rows)

            n      = len(df_trades)
            wins   = (df_trades['Profit_$'] > 0).sum()
            gl     = abs(df_trades.loc[df_trades['Profit_$'] < 0, 'Profit_$'].sum())
            gp     = df_trades.loc[df_trades['Profit_$'] > 0, 'Profit_$'].sum()
            pf     = round(gp / gl, 3) if gl > 0 else float('inf')
            longs  = df_trades[df_trades['Direction'] == 'LONG']
            shorts = df_trades[df_trades['Direction'] == 'SHORT']

            summary = pd.DataFrame([
                ['Total Trades',     n],
                ['Win Rate',         f"{wins/n*100:.1f}%" if n else '0%'],
                ['Profit Factor',    pf if pf != float('inf') else '∞'],
                ['Total Profit $',   round(df_trades['Profit_$'].sum(), 2)],
                ['LONG trades',      len(longs)],
                ['SHORT trades',     len(shorts)],
                ['LONG win rate',    f"{(longs['Profit_$']>0).mean()*100:.1f}%"
                                     if len(longs) else 'N/A'],
                ['SHORT win rate',   f"{(shorts['Profit_$']>0).mean()*100:.1f}%"
                                     if len(shorts) else 'N/A'],
                ['Timeframe',        getattr(self, 'timeframe', '1h')],
                ['Direction',        getattr(self, 'trade_direction', 'both')],
                ['Stop ATR Mult',    getattr(self, 'stop_loss_atr_mult', 2.2)],
                ['Quality Min L',    getattr(self, 'quality_min_long', 55)],
                ['Quality Min S',    getattr(self, 'quality_min_short', 57)],
            ], columns=['Metric', 'Value'])

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_trades.to_excel(writer, sheet_name='Trades',  index=False)
                summary.to_excel(writer,   sheet_name='Summary', index=False)
                wb = writer.book
                ws = wb['Trades']
                green_f = PatternFill("solid", fgColor="C6EFCE")
                red_f   = PatternFill("solid", fgColor="FFC7CE")
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    p_cell = next(
                        (c for c in row
                         if ws.cell(1, c.column).value == 'Profit_$'), None)
                    if p_cell is None:
                        continue
                    fill = green_f if (p_cell.value or 0) > 0 else red_f
                    for cell in row:
                        cell.fill = fill

            print(f"✅ Trades exported → {filename}  ({n} trades)")
            return True

        except Exception as e:
            import traceback
            print(f"❌ export_trades_to_excel: {e}")
            print(traceback.format_exc())
            return False

'''


# ════════════════════════════════════════════════════════════════════════════
# PATCH ROUTINES
# ════════════════════════════════════════════════════════════════════════════

def patch_app_macd(path: str) -> bool:
    hdr("Patching  App_MACD_AI_HybridScore_Latest1.py")
    p = FilePatcher(path)
    if not p.load():
        return False
    backup = p.backup()
    ok_count = err_count = skip_count = 0

    try:
        # ── Helper ─────────────────────────────────────────────────────────
        def apply(description, fn):
            nonlocal ok_count, err_count, skip_count
            result = fn()
            if result is True:
                ok_count += 1
            elif result == "skip":
                skip_count += 1
            else:
                err_count += 1

        # ── stop_trading ────────────────────────────────────────────────────
        def _patch_stop_trading():
            if p.has_method("stop_trading"):
                warn("stop_trading already exists — skipping"); return "skip"
            anchor = "    def start_trading(self):"
            if not p.has_anchor(anchor):
                err("Anchor for stop_trading not found"); return False
            p.insert_before(anchor, _STOP_TRADING,
                            "ADD stop_trading()")
            return True
        apply("ADD stop_trading", _patch_stop_trading)

        # ── predict_future_trend ────────────────────────────────────────────
        def _patch_predict():
            if p.has_method("predict_future_trend"):
                warn("predict_future_trend already exists — skipping"); return "skip"
            anchor = "    def stop_trading(self):"
            if not p.has_anchor(anchor):
                # fallback anchor
                anchor = "    def start_trading(self):"
            if not p.has_anchor(anchor):
                err("Anchor for predict_future_trend not found"); return False
            p.insert_before(anchor, _PREDICT_FUTURE,
                            "ADD predict_future_trend()")
            return True
        apply("ADD predict_future_trend", _patch_predict)

        # ── close_partial ───────────────────────────────────────────────────
        def _patch_close_partial():
            if p.has_method("close_partial"):
                warn("close_partial already exists — skipping"); return "skip"
            anchor = "    def stop_trading(self):"
            if not p.has_anchor(anchor):
                anchor = "    def start_trading(self):"
            if not p.has_anchor(anchor):
                err("Anchor for close_partial not found"); return False
            p.insert_before(anchor, _CLOSE_PARTIAL,
                            "ADD close_partial()")
            return True
        apply("ADD close_partial", _patch_close_partial)

        # ── Time-window methods ─────────────────────────────────────────────
        def _patch_time_window():
            if p.has_method("_init_trading_time_config"):
                warn("Time-window methods already exist — skipping"); return "skip"
            anchor = "    def stop_trading(self):"
            if not p.has_anchor(anchor):
                anchor = "    def start_trading(self):"
            if not p.has_anchor(anchor):
                err("Anchor for time-window methods not found"); return False
            p.insert_before(anchor, _TIME_WINDOW_METHODS,
                            "ADD Trading Time Window methods")
            return True
        apply("ADD Time Window methods", _patch_time_window)

        # ── Replace start_trading ───────────────────────────────────────────
        def _replace_start_trading():
            if not p.has_method("start_trading"):
                err("start_trading not found — cannot replace"); return False
            if "_is_time_unconstrained" in p.content:
                warn("start_trading already patched — skipping"); return "skip"
            result = p.replace_method("start_trading", _START_TRADING,
                                      description="MOD start_trading()")
            return result
        apply("MOD start_trading", _replace_start_trading)

        # ── __init__ time-config additions ──────────────────────────────────
        def _patch_init():
            if "_waiting_to_start" in p.content:
                warn("__init__ time-config already present — skipping"); return "skip"
            anchor = ("        self.weight_manager = "
                      "AdaptiveWeightManager(alpha=0.2, min_w=0.2, max_w=0.8)")
            if not p.has_anchor(anchor):
                err("__init__ anchor not found"); return False
            p.insert_after(anchor, _INIT_TIME_CONFIG,
                           "MOD __init__ — time-config lines")
            return True
        apply("MOD __init__", _patch_init)

        # ── trading_loop guard ───────────────────────────────────────────────
        def _patch_trading_loop():
            if "_strategy_name = self.strategy_type_var.get()" in p.content:
                warn("trading_loop guard already present — skipping"); return "skip"
            # Find the unique line inside trading_loop
            anchor = ("                df = self.get_market_data()\n"
                      "                if df is None:")
            if not p.has_anchor(anchor):
                err("trading_loop anchor not found"); return False
            insert_after_text = "                df = self.get_market_data()\n"
            p.insert_after(insert_after_text,
                           _TRADING_LOOP_PATCH,
                           "MOD trading_loop — time-window guard")
            return True
        apply("MOD trading_loop", _patch_trading_loop)

        # ── Settings tab hooks ───────────────────────────────────────────────
        def _patch_settings_hook(strategy, hook_code, anchor_method):
            marker = f'_add_time_settings_button(parent, "{strategy}")'
            if marker in p.content:
                warn(f"{strategy} settings hook already present — skipping")
                return "skip"
            # Find first line of the method body
            method_anchor = f"    def {anchor_method}(self, parent):\n"
            if not p.has_anchor(method_anchor):
                err(f"Anchor for {strategy} settings hook not found"); return False
            # Insert after the method definition line
            p.insert_after(method_anchor, hook_code,
                           f"ADD ⏰ hook to {anchor_method}")
            return True

        apply("ADD Momentum time hook",
              lambda: _patch_settings_hook("Momentum",
                                           _MOMENTUM_TIME_HOOK,
                                           "create_momentum_parameter_controls"))
        apply("ADD Kalman time hook",
              lambda: _patch_settings_hook("Kalman",
                                           _KALMAN_TIME_HOOK,
                                           "create_kalman_parameter_controls"))
        apply("ADD Scalping time hook",
              lambda: _patch_settings_hook("Scalping",
                                           _SCALPING_TIME_HOOK,
                                           "create_scalping_parameter_controls"))

        # ── Save ─────────────────────────────────────────────────────────────
        if err_count == 0:
            p.save()
            print(f"\n  {GREEN}{BOLD}App_MACD_AI_HybridScore_Latest1.py: "
                  f"{ok_count} applied, {skip_count} skipped, {err_count} errors{RESET}")
            return True
        else:
            p.rollback()
            print(f"\n  {RED}{BOLD}App_MACD_AI_HybridScore_Latest1.py: ROLLED BACK — "
                  f"{err_count} errors{RESET}")
            return False

    except Exception as exc:
        p.rollback()
        err(f"Unexpected error — rolled back: {exc}")
        import traceback; traceback.print_exc()
        return False


def patch_scalping_strategy(path: str) -> bool:
    hdr("Patching  scalping_strategy.py")
    p = FilePatcher(path)
    if not p.load():
        return False
    p.backup()
    ok_count = err_count = skip_count = 0

    try:
        def apply(description, fn):
            nonlocal ok_count, err_count, skip_count
            result = fn()
            if result is True:
                ok_count += 1
            elif result == "skip":
                skip_count += 1
            else:
                err_count += 1

        # ── ScalpingStrategy: execute_buy ────────────────────────────────────
        def _buy():
            if p.has_method("execute_buy"):
                warn("execute_buy already exists — skipping"); return "skip"
            anchor = "    def run_analysis_cycle(self, current_data, current_price, df=None):"
            if not p.has_anchor(anchor):
                err("Anchor for execute_buy not found"); return False
            p.insert_before(anchor, _SCALPING_EXECUTE_BUY, "ADD execute_buy()")
            return True
        apply("ADD execute_buy", _buy)

        # ── ScalpingStrategy: execute_sell ───────────────────────────────────
        def _sell():
            if p.has_method("execute_sell"):
                warn("execute_sell already exists — skipping"); return "skip"
            anchor = "    def run_analysis_cycle(self, current_data, current_price, df=None):"
            if not p.has_anchor(anchor):
                err("Anchor for execute_sell not found"); return False
            p.insert_before(anchor, _SCALPING_EXECUTE_SELL, "ADD execute_sell()")
            return True
        apply("ADD execute_sell", _sell)

        # ── ScalpingStrategy: get_strategy_info ─────────────────────────────
        def _info():
            if p.has_method("get_strategy_info"):
                warn("get_strategy_info already exists — skipping"); return "skip"
            anchor = "    def run_analysis_cycle(self, current_data, current_price, df=None):"
            if not p.has_anchor(anchor):
                err("Anchor for get_strategy_info not found"); return False
            p.insert_before(anchor, _SCALPING_GET_STRATEGY_INFO,
                            "ADD get_strategy_info()")
            return True
        apply("ADD get_strategy_info", _info)

        # ── BacktestScalpingStrategy: export_trades_to_excel ─────────────────
        def _export():
            if p.has_method("export_trades_to_excel"):
                warn("export_trades_to_excel already exists — skipping"); return "skip"
            # Insert before the end of BacktestScalpingStrategy (before next top-level class)
            anchor = "\nclass TimeframeManager:"
            if not p.has_anchor(anchor):
                # fallback: insert before the GUI integration comment block
                anchor = "\n# ═══════════════════════════════════════════════════════════════════════════\n# GUI TIMEFRAME INTEGRATION"
            if not p.has_anchor(anchor):
                err("Anchor for export_trades_to_excel not found"); return False
            p.insert_before(anchor, _BACKTEST_EXPORT_TRADES,
                            "ADD export_trades_to_excel()")
            return True
        apply("ADD export_trades_to_excel", _export)

        if err_count == 0:
            p.save()
            print(f"\n  {GREEN}{BOLD}scalping_strategy.py: "
                  f"{ok_count} applied, {skip_count} skipped, {err_count} errors{RESET}")
            return True
        else:
            p.rollback()
            print(f"\n  {RED}{BOLD}scalping_strategy.py: ROLLED BACK — "
                  f"{err_count} errors{RESET}")
            return False

    except Exception as exc:
        p.rollback()
        err(f"Unexpected error — rolled back: {exc}")
        import traceback; traceback.print_exc()
        return False


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def find_file(candidates: list) -> str | None:
    for c in candidates:
        if Path(c).exists():
            return c
    return None


def main():
    print(f"\n{BOLD}{CYAN}"
          f"╔══════════════════════════════════════════════════════════════════╗\n"
          f"║   TRADING APP — UPDATE PACKAGE  v2.0                            ║\n"
          f"║   {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}                                    ║\n"
          f"╚══════════════════════════════════════════════════════════════════╝"
          f"{RESET}\n")

    # ── Locate source files ───────────────────────────────────────────────
    app_path = find_file([
        "App_MACD_AI_HybridScore_Latest1.py",
        "app/App_MACD_AI_HybridScore_Latest1.py",
        "src/App_MACD_AI_HybridScore_Latest1.py",
    ])
    sc_path = find_file([
        "scalping_strategy.py",
        "strategies/scalping_strategy.py",
        "src/strategies/scalping_strategy.py",
    ])

    if app_path is None:
        err("App_MACD_AI_HybridScore_Latest1.py not found. Run this script from your project root.")
        sys.exit(1)
    if sc_path is None:
        err("scalping_strategy.py not found. Run this script from your project root.")
        sys.exit(1)

    info(f"App_MACD_AI_HybridScore_Latest1.py       → {app_path}")
    info(f"scalping_strategy → {sc_path}")

    # ── Apply patches ─────────────────────────────────────────────────────
    results = {
        "App_MACD_AI_HybridScore_Latest1.py":          patch_app_macd(app_path),
        "scalping_strategy.py": patch_scalping_strategy(sc_path),
    }

    # ── Summary ───────────────────────────────────────────────────────────
    hdr("UPDATE SUMMARY")
    all_ok = True
    for filename, success in results.items():
        if success:
            ok(f"{filename}  — PATCHED SUCCESSFULLY")
        else:
            err(f"{filename}  — FAILED  (original restored from backup)")
            all_ok = False

    if all_ok:
        print(f"\n{BOLD}{GREEN}"
              f"  ✅  All patches applied.  You can now run your application.\n"
              f"{RESET}")
    else:
        print(f"\n{BOLD}{RED}"
              f"  ❌  One or more patches failed.  Originals have been restored.\n"
              f"      Check the error messages above and re-run after fixing.\n"
              f"{RESET}")
        sys.exit(1)


if __name__ == "__main__":
    main()