"""
╔══════════════════════════════════════════════════════════════════════════╗
║   PROFESSIONAL SCALPING STRATEGY v1.7.3 — SMART PARTIAL EXITS          ║
║   Target: 30-50 trades/month | WR >47% | PF >1.3 | Max DD <10%         ║
╠══════════════════════════════════════════════════════════════════════════╣
║  v1.7.3 FIXES:                                                          ║
║  1. INTELLIGENT PARTIAL EXIT DECISION LOGIC                            ║
║     - Evaluates EV of remaining position before allowing partial       ║
║     - Prevents partial exits when remaining volume has negative EV    ║
║     - Fixed cost ratio check (max 5% of edge)                         ║
║     - Alpha decay protection (30% edge decay on remaining)            ║
║     - Minimum remaining volume check (30% of original)               ║
║                                                                          ║
║  v1.7.2 FIXES:                                                          ║
║  1. REMOVED PARTIAL EXITS — full exits only, saves ~$169 commission    ║
║  2. RAISED QUALITY THRESHOLDS — Tier1 80L/80S, Tier2 72L/72S           ║
║  3. MAKER ORDER SUPPORT — 0.04% round trip vs 0.20%, saves $1,680+    ║
║                                                                          ║
║  NOTE: No trading strategy is guaranteed profitable. This tightens      ║
║  risk management (tier/direction-aware sizing & stops) versus the       ║
║  prior single-tier version; validate with walk-forward testing on       ║
║  your own data before risking live capital.                            ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import json
import logging
import os
import math
from datetime import datetime, timedelta, timezone
from dataclasses import dataclass
from enum import Enum, auto
from typing import Dict, List, Tuple, Optional, Any
from collections import deque
import numpy as np
import pandas as pd
import talib
from .base3_New import BaseStrategy
from backtesting import Strategy


# ═══════════════════════════════════════════════════════════════════════════
# TRADE DIRECTION NORMALIZATION
# ═══════════════════════════════════════════════════════════════════════════

def _normalize_trade_direction(value, default="both"):
    if value is None:
        return default
    v = str(value).strip().lower()
    return v if v in ('long', 'short', 'both') else default


# ═══════════════════════════════════════════════════════════════════════════
# DIRECTION MANAGER
# ═══════════════════════════════════════════════════════════════════════════

class DirectionManager:
    _instance = None
    _direction = "both"
    _initialized = False

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def initialize(self, direction: str = "both"):
        norm = _normalize_trade_direction(direction)
        self._direction = norm
        self._initialized = True
        os.environ['SCALPING_DIRECTION_MANAGER_STATE'] = norm
        print(f"🎯 DIRECTION MANAGER INITIALIZED: {norm.upper()}")

    def set_direction(self, direction: str):
        norm = _normalize_trade_direction(direction, default=None)
        if norm is None:
            print(f"⚠️ DIRECTION MANAGER: Invalid direction '{direction}', keeping {self._direction}")
            return
        if not self._initialized:
            self._direction = norm
            self._initialized = True
            os.environ['SCALPING_DIRECTION_MANAGER_STATE'] = norm
            print(f"🎯 DIRECTION MANAGER INITIALIZED: {norm.upper()}")
            return
        if self._direction != norm:
            self._direction = norm
            os.environ['SCALPING_DIRECTION_MANAGER_STATE'] = norm
            print(f"🎯 DIRECTION MANAGER: Updated to {norm.upper()}")

    def get_direction(self) -> str:
        if not self._initialized:
            persisted = os.environ.get('SCALPING_DIRECTION_MANAGER_STATE')
            fallback = _normalize_trade_direction(
                persisted, default=SCALPING_PARAMS.get('trade_direction', 'both'))
            self.initialize(fallback)
        return self._direction

    def is_long_allowed(self) -> bool:
        return self.get_direction() in ['long', 'both']

    def is_short_allowed(self) -> bool:
        return self.get_direction() in ['short', 'both']

    def is_trade_allowed(self, direction: str) -> bool:
        if direction == 'long':
            return self.is_long_allowed()
        elif direction == 'short':
            return self.is_short_allowed()
        return False

    def get_direction_display(self) -> str:
        d = self.get_direction()
        return d.upper()

    def __repr__(self):
        return f"DirectionManager(direction={self._direction.upper()}, initialized={self._initialized})"


DIRECTION_MANAGER = DirectionManager()


# ═══════════════════════════════════════════════════════════════════════════
# GLOBAL CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════

class GlobalConfig:
    INITIAL_CAPITAL = 50000.0
    COMMISSION_RATE = 0.0005
    DEFAULT_SYMBOL = "SOL-USDT"
    ACTIVE_TIMEFRAME = "1h"
    VALID_TIMEFRAMES = ["1m", "5m", "15m", "30m", "1h", "2h", "4h", "1d"]
    TIMEFRAME_MINUTES = {"1m": 1, "5m": 5, "15m": 15, "30m": 30, "1h": 60, "2h": 120, "4h": 240, "1d": 1440}
    TIMEFRAME_MULTIPLIERS = {"1m": 1 / 60., "5m": 5 / 60., "15m": .25, "30m": .5, "1h": 1., "2h": 2., "4h": 4.,
                             "1d": 24.}

    @classmethod
    def update_capital(cls, v):
        old = cls.INITIAL_CAPITAL
        cls.INITIAL_CAPITAL = float(v)
        logging.info(f"💰 CAPITAL ${old:,.0f}→${cls.INITIAL_CAPITAL:,.0f}")
        return cls.INITIAL_CAPITAL

    @classmethod
    def update_timeframe(cls, tf):
        if tf in cls.VALID_TIMEFRAMES:
            cls.ACTIVE_TIMEFRAME = tf
        return cls.ACTIVE_TIMEFRAME

    @classmethod
    def get_bar_interval_minutes(cls):
        return cls.TIMEFRAME_MINUTES.get(cls.ACTIVE_TIMEFRAME, 60)

    @classmethod
    def get_timeframe_multiplier(cls):
        return cls.TIMEFRAME_MULTIPLIERS.get(cls.ACTIVE_TIMEFRAME, 1.)


CAPITAL = GlobalConfig.INITIAL_CAPITAL


# ═══════════════════════════════════════════════════════════════════════════
# TIMEFRAME SCALING
# ═══════════════════════════════════════════════════════════════════════════

def scale_params_for_timeframe(params: dict, timeframe: str) -> dict:
    from copy import deepcopy
    sc = deepcopy(params)
    mult = GlobalConfig.TIMEFRAME_MULTIPLIERS.get(timeframe, 1.0)

    pullback_was_default = (
            sc.get('pullback_zone_lower_pct') == SCALPING_PARAMS.get('pullback_zone_lower_pct') and
            sc.get('pullback_zone_upper_pct') == SCALPING_PARAMS.get('pullback_zone_upper_pct'))
    quality_was_default = (
            sc.get('quality_min_long') == SCALPING_PARAMS.get('quality_min_long') and
            sc.get('quality_min_short') == SCALPING_PARAMS.get('quality_min_short') and
            sc.get('quality_tier1_min_long') == SCALPING_PARAMS.get('quality_tier1_min_long') and
            sc.get('quality_tier1_min_short') == SCALPING_PARAMS.get('quality_tier1_min_short'))
    volume_min_ratio_was_default = sc.get('volume_min_ratio') == SCALPING_PARAMS.get('volume_min_ratio')
    volume_strong_ratio_was_default = sc.get('volume_strong_ratio') == SCALPING_PARAMS.get('volume_strong_ratio')

    linear = ['ema_fast_period', 'ema_mid_period', 'ema_slow_period',
              'macd_fast', 'macd_slow', 'macd_signal_period', 'atr_period', 'rsi_period', 'adx_period',
              'volume_period', 'daily_ema_period', 'max_hold_bars', 'min_hold_bars_before_stop',
              'min_bars_between_trades', 'cooldown_after_loss_bars', 'consecutive_loss_cooldown_bars',
              'trend_age_max_bars', 'extended_run_lookback', 'atr_compression_lookback',
              'be_stop_no_progress_bars', 'momentum_period', 'chop_period', 'bb_period', 'kc_period']
    floors = {'ema_fast_period': 5, 'ema_mid_period': 10, 'ema_slow_period': 20,
              'macd_fast': 5, 'macd_slow': 12, 'macd_signal_period': 3,
              'atr_period': 7, 'rsi_period': 7, 'adx_period': 7,
              'volume_period': 10, 'chop_period': 10, 'bb_period': 10, 'kc_period': 10, 'momentum_period': 1}
    for p in linear:
        if p in sc:
            orig = sc[p]
            sc[p] = max(max(1, int(math.ceil(orig * mult))), floors.get(p, 1))

    for p in ['stop_loss_atr_mult', 'emergency_stop_mult', 'trailing_activation_pct',
              'trailing_distance_pct', 'pullback_zone_lower_pct', 'pullback_zone_upper_pct',
              'tier1_stop_mult_long', 'tier1_stop_mult_short',
              'tier2_stop_mult_long', 'tier2_stop_mult_short',
              'trailing_activation_tier1', 'trailing_activation_tier2',
              'trailing_distance_tier1', 'trailing_distance_tier2']:
        if p in sc:
            orig = sc[p]
            sc[p] = orig * (max(0.3, mult) if mult < 1 else min(3., mult))

    sc['daily_ema_period'] = max(20, sc.get('daily_ema_period', 24))

    bm = GlobalConfig.TIMEFRAME_MINUTES.get(timeframe, 60)
    if pullback_was_default:
        if bm <= 5:
            sc['pullback_zone_lower_pct'] = -0.8
            sc['pullback_zone_upper_pct'] = 0.5
        elif bm <= 15:
            sc['pullback_zone_lower_pct'] = -3.0
            sc['pullback_zone_upper_pct'] = 1.5

    if quality_was_default:
        if mult <= 5 / 60:
            sc['quality_min_long'] = 70
            sc['quality_min_short'] = 70
            sc['quality_tier1_min_long'] = 80
            sc['quality_tier1_min_short'] = 80
            sc['quality_tier2_min_long'] = 70
            sc['quality_tier2_min_short'] = 70
        elif mult < 0.5:
            sc['quality_min_long'] = 72
            sc['quality_min_short'] = 72
            sc['quality_tier1_min_long'] = 82
            sc['quality_tier1_min_short'] = 82
            sc['quality_tier2_min_long'] = 72
            sc['quality_tier2_min_short'] = 72

    if mult < 0.5:
        if volume_min_ratio_was_default:
            sc['volume_min_ratio'] = max(0.8, SCALPING_PARAMS.get('volume_min_ratio', 0.9) * 0.9)
        if volume_strong_ratio_was_default:
            sc['volume_strong_ratio'] = max(1.3, SCALPING_PARAMS.get('volume_strong_ratio', 1.6) * 0.9)

    sc['bar_interval_minutes'] = bm
    sc['timeframe'] = timeframe
    return sc


def get_indicator_periods_for_timeframe(timeframe: str) -> dict:
    base = {'ema_fast': 9, 'ema_mid': 21, 'ema_slow': 50, 'macd_fast': 12, 'macd_slow': 26, 'macd_signal': 9,
            'rsi': 14, 'atr': 14, 'adx': 14, 'stoch_k': 5, 'stoch_d': 3}
    mult = GlobalConfig.TIMEFRAME_MULTIPLIERS.get(timeframe, 1.)
    if timeframe in ['1m', '5m']:
        es, ms = max(.15, mult * .8), max(.2, mult)
    elif timeframe in ['15m', '30m']:
        es, ms = mult * .9, mult * .95
    elif timeframe in ['2h', '4h']:
        es, ms = min(1.8, mult * .9), min(1.8, mult * .9)
    elif timeframe == '1d':
        es, ms = min(3.5, mult * .7), min(3., mult * .7)
    else:
        es, ms = mult, mult
    return {'ema_fast_period': max(3, int(base['ema_fast'] * es)),
            'ema_mid_period': max(5, int(base['ema_mid'] * es)),
            'ema_slow_period': max(10, int(base['ema_slow'] * es)),
            'macd_fast': max(3, int(base['macd_fast'] * ms)),
            'macd_slow': max(5, int(base['macd_slow'] * ms)),
            'macd_signal_period': max(3, int(base['macd_signal'] * ms)),
            'rsi_period': max(7, int(base['rsi'] * max(.7, min(1.3, mult)))),
            'atr_period': max(7, int(base['atr'] * max(.7, min(1.3, mult)))),
            'adx_period': max(7, int(base['adx'] * max(.7, min(1.3, mult))))}


# ═══════════════════════════════════════════════════════════════════════════
# UTILITIES
# ═══════════════════════════════════════════════════════════════════════════

def safe_profit_factor(gp, gl):
    return float('inf') if gl <= 0 else gp / gl


def compute_sharpe(returns, risk_free=0., periods_per_year=None):
    if len(returns) < 2:
        return 0.
    if periods_per_year is None:
        periods_per_year = int((1440 / max(GlobalConfig.get_bar_interval_minutes(), 1)) * 252)
    s = np.std(returns)
    return ((np.mean(returns) - risk_free) / s) * np.sqrt(periods_per_year) if s > 0 else 0.


def compute_sortino(returns, target=0., periods_per_year=None):
    if len(returns) < 2:
        return 0.
    if periods_per_year is None:
        periods_per_year = int((1440 / max(GlobalConfig.get_bar_interval_minutes(), 1)) * 252)
    down = returns[returns < target]
    if len(down) == 0:
        return float('inf')
    d = np.sqrt(np.mean(down ** 2))
    return (np.mean(returns) / d) * np.sqrt(periods_per_year) if d > 0 else float('inf')


def summarize_performance(trades, initial_capital=None):
    if initial_capital is None:
        initial_capital = GlobalConfig.INITIAL_CAPITAL
    n = len(trades)
    if n == 0:
        return {"total_trades": 0, "win_rate": 0., "profit_pct": 0.,
                "profit_factor": None, "expectancy": 0., "warning": "No trades"}
    tp = sum(getattr(t, 'profit', t.get('profit', 0)) for t in trades)
    wins = sum(1 for t in trades if getattr(t, 'profit', t.get('profit', 0)) > 0)
    losses = n - wins
    gp = sum(getattr(t, 'profit', t.get('profit', 0)) for t in trades if getattr(t, 'profit', t.get('profit', 0)) > 0)
    gl = -sum(getattr(t, 'profit', t.get('profit', 0)) for t in trades if getattr(t, 'profit', t.get('profit', 0)) < 0)
    wr = wins / n
    aw = gp / wins if wins > 0 else 0.
    al = gl / losses if losses > 0 else 0.
    return {"total_trades": n, "win_rate": round(wr, 4), "profit_pct": round(tp / initial_capital * 100, 4),
            "profit_factor": round(safe_profit_factor(gp, gl), 4),
            "expectancy": round(wr * aw - (1 - wr) * al, 6),
            "avg_profit_per_trade": round(tp / n, 6), "gross_profit": round(gp, 2), "gross_loss": round(gl, 2),
            "wins": wins, "losses": losses,
            "warning": f"⚠️ Small sample ({n})" if n < 30 else ""}


# ═══════════════════════════════════════════════════════════════════════════
# STATE MACHINE
# ═══════════════════════════════════════════════════════════════════════════

class StrategyState(Enum):
    SEEKING_ENTRY = auto()
    IN_TRADE = auto()


POSITION_ALREADY_OPEN_SENTINEL = -1


# ═══════════════════════════════════════════════════════════════════════════
# DATA CLASSES
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class TradeRecord:
    trade_id: int
    symbol: str
    entry_time: datetime
    entry_price: float
    entry_size: float
    entry_tier: int
    entry_quality_score: int
    entry_reason: str
    entry_direction: str
    exit_time: Optional[datetime] = None
    exit_price: Optional[float] = None
    exit_size: Optional[float] = None
    exit_reason: Optional[str] = None
    profit: float = 0.
    profit_pct: float = 0.
    profit_r: float = 0.
    max_profit: float = 0.
    max_drawdown_pct: float = 0.
    hold_duration: float = 0.
    market_regime: str = "UNKNOWN"
    partial_exits_taken: int = 0
    partial_pnl_realised: float = 0.
    original_size: float = 0.

    def to_dict(self):
        return {k: v for k, v in self.__dict__.items()}


@dataclass
class RiskMetrics:
    daily_loss: float = 0.
    max_drawdown: float = 0.
    consecutive_losses: int = 0
    win_rate: float = 0.
    profit_factor: float = 0.
    sharpe_ratio: float = 0.
    sortino_ratio: float = 0.
    expectancy: float = 0.


# ═══════════════════════════════════════════════════════════════════════════
# SCALPING PARAMETERS — v1.7.3 (Smart Partial Exits)
# ═══════════════════════════════════════════════════════════════════════════

SCALPING_PARAMS = {
    "trade_direction": "short",
    "ema_fast_period": 9, "ema_mid_period": 21, "ema_slow_period": 50,
    "macd_fast": 12, "macd_slow": 26, "macd_signal_period": 9,
    "stoch_k_period": 5, "stoch_d_period": 3, "stoch_smooth": 3,
    "stoch_overbought": 80, "stoch_oversold": 20,
    "stoch_mid_upper": 70, "stoch_mid_lower": 30,
    "rsi_period": 14,
    "rsi_long_min": 40, "rsi_long_max": 70,
    "rsi_short_min": 30, "rsi_short_max": 60,
    "rsi_overbought_exit": 75, "rsi_oversold_exit": 25,
    "adx_period": 14,
    "adx_min_long": 22, "adx_min_short": 22,
    "adx_extended_threshold": 55,
    "volume_period": 20,
    "volume_min_ratio": 0.9, "volume_strong_ratio": 1.6,
    "atr_period": 14, "atr_compression_lookback": 40, "atr_compression_threshold": 0.35,
    "atr_spike_ratio": 1.8,

    "quality_min_long": 72,
    "quality_min_short": 72,
    "quality_tier1_min_long": 80,
    "quality_tier1_min_short": 80,
    "quality_tier2_min_long": 72,
    "quality_tier2_min_short": 72,

    "commission_round_trip_pct": 0.0010, "early_exit_min_profit_multiple": 2.5,
    "weight_ema": 35, "weight_macd": 20, "weight_rsi": 15,
    "weight_volume": 10, "weight_adx": 10, "weight_price_ema": 10,
    "risk_per_trade": 0.008,
    "tier1_risk_pct": 0.010, "tier2_risk_pct": 0.006,
    "tier1_size_mult": 1.00, "tier2_size_mult": 0.65,
    "max_position_size_pct": 0.15, "max_position_units": 100,
    "min_cash_reserve": 0.20, "base_risk_pct": 0.008,
    "stop_loss_atr_mult": 2.2, "emergency_stop_mult": 3.5,
    "tier1_stop_mult_long": 1.8, "tier1_stop_mult_short": 2.4,
    "tier2_stop_mult_long": 2.2, "tier2_stop_mult_short": 2.8,
    "trailing_activation_pct": 0.020, "trailing_distance_pct": 0.012, "trailing_atr_mult": 1.0,
    "trailing_activation_tier1": 0.016, "trailing_activation_tier2": 0.024,
    "trailing_distance_tier1": 0.010, "trailing_distance_tier2": 0.014,

    "be_stop_enabled": True, "be_stop_r_trigger": 1.5, "be_stop_no_progress_bars": 10,
    "take_profit_r1": 1.5,
    "take_profit_r2": 2.8,
    "min_take_profit_r": 2.0,
    "full_exit_at_r2": True,
    "partial_exit_pct_r1": 0.50,
    "partial_exit_pct_r2": 0.30,

    "macd_cross_exit_enabled": True, "macd_cross_min_profit_r": 0.8,
    "stoch_reversal_exit_enabled": True, "stoch_reversal_min_profit_r": 0.5,
    "ema_cross_exit_enabled": True, "ema_cross_min_profit_r": 1.5,
    "runner_ema_reversal_exit_enabled": True, "runner_ema_reversal_min_profit_r": 0.3,
    "max_hold_bars": 24, "min_hold_bars_before_stop": 2,
    "pullback_zone_lower_pct": -3.0, "pullback_zone_upper_pct": 1.5,
    "adx_slope_min": -0.5,
    "momentum_period": 3, "momentum_min_long": 0.003, "momentum_min_short": 0.003,
    "trend_age_min_bars": 3,
    "regime_filter_enabled": True,
    "bb_period": 20, "bb_std": 2.0, "kc_period": 20, "kc_atr_mult": 1.5,
    "chop_period": 14, "chop_threshold": 61, "ranging_min_checks": 4,
    "max_daily_trades": 10, "min_bars_between_trades": 4,
    "cooldown_after_loss_bars": 5,
    "consecutive_loss_threshold": 3, "consecutive_loss_cooldown_bars": 5,
    "daily_trend_filter_enabled": True, "daily_ema_period": 24, "daily_trend_adx_override": 20,
    "extended_run_lookback": 6, "extended_run_max_pct_long": 5.0, "extended_run_max_pct_short": 5.0,
    "trend_age_penalty_enabled": True, "trend_age_max_bars": 24, "trend_age_penalty_pts": 5,
    "daily_loss_limit_pct": 0.02, "max_drawdown_limit_pct": 0.10, "max_consecutive_losses": 5,
    "only_tier2_entries": False,
    "bar_interval_minutes": None, "timeframe": "1h",
    "fuzzy_mode_enabled": False, "fuzzy_learning_enabled": True,
    "fuzzy_absolute_min": 45, "fuzzy_absolute_max": 65,
    "fuzzy_default_margin_pct": 10, "fuzzy_min_confidence": 0.60, "fuzzy_min_samples": 8,

    "use_maker_orders": True,
    "maker_order_offset_pct": 0.001,
    "maker_order_timeout_bars": 3,
    "force_maker_rebate": True,

    # NEW: Smart Partial Exit Control
    "allow_partial_exits": False,
    "partial_exit_min_rem_edge_pct": 0.15,
    "partial_exit_fixed_cost_ratio": 0.05,
    "partial_exit_min_volume_pct": 0.30,
    "partial_exit_alpha_decay_pct": 0.30,
    "partial_exit_max_legs": 2,
    "partial_exit_cancel_residual": True,
    "debug_partial_exit": False,
}


# ═══════════════════════════════════════════════════════════════════════════
# RISK CONTROLLER
# ═══════════════════════════════════════════════════════════════════════════

class ScalpingRiskController:
    def __init__(self, starting_equity=None):
        if starting_equity is None:
            starting_equity = GlobalConfig.INITIAL_CAPITAL
        self.starting_equity = starting_equity
        self.current_equity = starting_equity
        self.peak_equity = starting_equity
        self.daily_loss_limit = starting_equity * SCALPING_PARAMS['daily_loss_limit_pct']
        self.max_drawdown_limit = starting_equity * SCALPING_PARAMS['max_drawdown_limit_pct']
        self.max_consecutive = SCALPING_PARAMS['max_consecutive_losses']
        self.max_position_pct = SCALPING_PARAMS['max_position_size_pct']
        self.max_position_units = SCALPING_PARAMS['max_position_units']
        self.min_cash_reserve = SCALPING_PARAMS['min_cash_reserve']
        self.base_risk_pct = SCALPING_PARAMS['base_risk_pct']

        self.tier1_risk_pct = SCALPING_PARAMS['tier1_risk_pct']
        self.tier2_risk_pct = SCALPING_PARAMS['tier2_risk_pct']
        self.tier1_size_mult = SCALPING_PARAMS['tier1_size_mult']
        self.tier2_size_mult = SCALPING_PARAMS['tier2_size_mult']
        self.tier1_stop_mult_long = SCALPING_PARAMS['tier1_stop_mult_long']
        self.tier1_stop_mult_short = SCALPING_PARAMS['tier1_stop_mult_short']
        self.tier2_stop_mult_long = SCALPING_PARAMS['tier2_stop_mult_long']
        self.tier2_stop_mult_short = SCALPING_PARAMS['tier2_stop_mult_short']

        self.trades: List[TradeRecord] = []
        self.today_loss = 0.
        self.today_date = datetime.now().date()
        self.equity_curve = [starting_equity]
        self.consecutive_losses = 0
        self.total_trades = 0
        self.winning_trades = 0
        self.losing_trades = 0
        self.risk_metrics = RiskMetrics()

    def get_tier_config(self, tier: int, direction: str = 'long') -> Optional[dict]:
        if tier == 1:
            stop_mult = self.tier1_stop_mult_short if direction == 'short' else self.tier1_stop_mult_long
            return {'risk_pct': self.tier1_risk_pct, 'size_mult': self.tier1_size_mult,
                    'stop_mult': stop_mult, 'name': 'Tier 1 (High Conviction)'}
        elif tier == 2:
            stop_mult = self.tier2_stop_mult_short if direction == 'short' else self.tier2_stop_mult_long
            return {'risk_pct': self.tier2_risk_pct, 'size_mult': self.tier2_size_mult,
                    'stop_mult': stop_mult, 'name': 'Tier 2 (Standard)'}
        return None

    def calculate_position_size(self, entry_price, stop_loss_price, quality_score=70,
                                adx=25., tier=1, direction='long'):
        if entry_price <= 0:
            return 0
        rp = abs(entry_price - stop_loss_price) / entry_price
        if rp <= 0 or rp > 0.15:
            return 0

        tier_cfg = self.get_tier_config(tier, direction) or self.get_tier_config(2, direction)
        tier_risk_pct = tier_cfg['risk_pct']
        tier_size_mult = tier_cfg['size_mult']

        qw = max(.6, min((quality_score / 70) ** .5, 1.4))
        am = (.6 if adx < 20 else .85 if adx < 25 else 1. if adx < 35 else .9 if adx < 45 else .6)
        sm = .7 if self.consecutive_losses >= 2 else 1.
        h = min((self.current_equity / max(1., self.peak_equity)) ** 2, 1.)
        rp2 = max(.002, min(tier_risk_pct * qw * am * sm * h, .025))
        sz = min(self.current_equity * rp2 / (entry_price * rp),
                 self.current_equity * self.max_position_pct / entry_price,
                 self.max_position_units) * tier_size_mult
        if entry_price >= 1000:
            return max(0., round(sz, 6))
        elif entry_price >= 100:
            return max(0., round(sz, 4))
        else:
            return max(0, int(sz))

    def validate_entry(self, position_size, entry_price):
        if self.today_loss <= -self.daily_loss_limit:
            return False, "daily_loss_limit"
        dd = (self.peak_equity - self.current_equity) / self.peak_equity
        if dd >= self.max_drawdown_limit:
            return False, f"max_dd_{dd:.1%}"
        if self.consecutive_losses >= self.max_consecutive:
            return False, "consec_loss"
        if position_size * entry_price > self.current_equity * (1 - self.min_cash_reserve):
            return False, "no_cash"
        if (position_size * entry_price) / self.current_equity > self.max_position_pct:
            return False, "too_large"
        return True, "pass"

    def record_trade(self, trade):
        self.trades.append(trade)
        self.total_trades += 1
        if trade.profit > 0:
            self.winning_trades += 1
            self.consecutive_losses = 0
        else:
            self.losing_trades += 1
            self.consecutive_losses += 1
        if trade.exit_time and trade.exit_time.date() == self.today_date:
            self.today_loss += trade.profit
        self.current_equity += trade.profit
        if self.current_equity > self.peak_equity:
            self.peak_equity = self.current_equity
        self.equity_curve.append(self.current_equity)
        self._recalculate_metrics()

    def _recalculate_metrics(self):
        if not self.total_trades:
            return
        self.risk_metrics.win_rate = self.winning_trades / self.total_trades
        wins = sum(t.profit for t in self.trades if t.profit > 0)
        losses = abs(sum(t.profit for t in self.trades if t.profit < 0))
        self.risk_metrics.profit_factor = safe_profit_factor(wins, losses)
        pk = self.starting_equity
        for eq in self.equity_curve:
            if eq > pk:
                pk = eq
            dd = (pk - eq) / pk
            if dd > self.risk_metrics.max_drawdown:
                self.risk_metrics.max_drawdown = dd
        self.risk_metrics.expectancy = sum(t.profit for t in self.trades) / self.total_trades
        if len(self.equity_curve) > 1:
            rets = np.diff(self.equity_curve) / np.array(self.equity_curve[:-1])
            if np.std(rets) > 0:
                self.risk_metrics.sharpe_ratio = compute_sharpe(rets)
                self.risk_metrics.sortino_ratio = compute_sortino(rets)

    def get_stats(self):
        rm = self.risk_metrics
        return {'total_trades': self.total_trades,
                'win_rate': f"{rm.win_rate * 100:.1f}%",
                'profit_factor': f"{rm.profit_factor:.2f}" if rm.profit_factor != float('inf') else "∞",
                'max_drawdown': f"{rm.max_drawdown:.1%}",
                'sharpe_ratio': f"{rm.sharpe_ratio:.2f}",
                'sortino_ratio': f"{rm.sortino_ratio:.2f}",
                'expectancy': f"${rm.expectancy:.2f}",
                'current_equity': f"${self.current_equity:,.2f}",
                'total_profit': f"${self.current_equity - self.starting_equity:,.2f}",
                'roi': f"{(self.current_equity - self.starting_equity) / self.starting_equity:.1%}",
                'consecutive_loss': self.consecutive_losses}


# ═══════════════════════════════════════════════════════════════════════════
# REGIME DETECTOR
# ═══════════════════════════════════════════════════════════════════════════

class RegimeDetector:
    def __init__(self):
        self.current_regime = "UNKNOWN"
        self.regime_confidence = 0.
        self.regime_history = deque(maxlen=100)
        self._atr_spike_ratio = SCALPING_PARAMS.get("atr_spike_ratio", 1.8)

    def detect_regime(self, ema_fast, ema_slow, adx, bb_width_pct=50,
                      atr_now: float = 0., atr_avg: float = 0.) -> tuple:
        if atr_now > 0 and atr_avg > 0:
            if atr_now > atr_avg * self._atr_spike_ratio:
                self.current_regime = "SPIKE"
                self.regime_confidence = 0.95
                self.regime_history.append("SPIKE")
                return "SPIKE", 0.95

        up = ema_fast > ema_slow
        strg = adx > 20
        weak = adx > 15
        sqz = bb_width_pct < 30

        if up and strg:
            r, c = "TRENDING_UP", 0.90
        elif not up and strg:
            r, c = "TRENDING_DOWN", 0.90
        elif up and weak:
            r, c = "TRENDING_UP_WEAK", 0.70
        elif not up and weak:
            r, c = "TRENDING_DOWN_WEAK", 0.70
        elif sqz:
            r, c = "RANGING_TIGHT", 0.80
        else:
            r, c = "CHOPPY", 0.55

        self.current_regime = r
        self.regime_confidence = c
        self.regime_history.append(r)
        return r, c

    def is_tradeable(self, regime, direction='both') -> bool:
        if regime in ("RANGING_TIGHT", "CHOPPY", "SPIKE"):
            return False
        if direction == 'long' and regime in ("TRENDING_DOWN", "TRENDING_DOWN_WEAK"):
            return False
        if direction == 'short' and regime in ("TRENDING_UP", "TRENDING_UP_WEAK"):
            return False
        return True

    def get_position_multiplier(self, regime) -> float:
        return {
            "TRENDING_UP": 1.2,
            "TRENDING_DOWN": 1.0,
            "TRENDING_UP_WEAK": 0.8,
            "TRENDING_DOWN_WEAK": 0.7,
            "RANGING_TIGHT": 0.4,
            "CHOPPY": 0.5,
            "SPIKE": 0.0,
        }.get(regime, 0.6)


# ═══════════════════════════════════════════════════════════════════════════
# EXIT MANAGER — v1.7.3 with Smart Partial Exits
# ═══════════════════════════════════════════════════════════════════════════

class ScalpingExitManager:
    def __init__(self, config):
        self.cfg = config
        self.full_exit_only = self.cfg.get('full_exit_at_r2', True)
        self.allow_partials = self.cfg.get('allow_partial_exits', False)
        self.debug = self.cfg.get('debug_partial_exit', False)

    def evaluate_partial_exit(self, current_price, entry_price, position_type,
                              original_quantity, remaining_quantity,
                              partial_exits_taken, profit_r, atr):
        """
        Evaluate whether a partial exit is profitable.

        Formula: EV_rem = (V_rem * E[R]) - (C_fixed + V_rem * C_var)

        Returns: (bool, reason_string)
        """
        # Don't allow partials if disabled
        if not self.allow_partials:
            return False, "partial_exits_disabled"

        # Maximum 2 partial legs
        max_legs = self.cfg.get('partial_exit_max_legs', 2)
        if partial_exits_taken >= max_legs:
            return False, "max_legs_reached"

        # Calculate volumes
        v_total = original_quantity
        v_rem = remaining_quantity

        if v_rem <= 0:
            return False, "no_remaining_volume"

        # Estimate edge per unit (using R-multiple and ATR)
        edge_per_unit = profit_r * atr

        # Fixed costs (commission per order)
        c_fixed = v_total * entry_price * self.cfg.get('commission_round_trip_pct', 0.001)
        c_var = self.cfg.get('commission_round_trip_pct', 0.001)

        # Alpha decay factor - remaining position has lower expected edge
        alpha_decay = self.cfg.get('partial_exit_alpha_decay_pct', 0.30)
        decayed_edge = edge_per_unit * (1 - alpha_decay)

        # Calculate EV of remaining position
        ev_rem = (v_rem * decayed_edge) - (c_fixed + v_rem * entry_price * c_var)

        # Cost comparison: single vs partial execution
        cost_single = c_fixed + (v_total * entry_price * c_var)
        cost_partial = (2 * c_fixed) + (v_total * entry_price * c_var)

        # Fixed cost ratio
        total_edge = v_total * abs(edge_per_unit)
        if total_edge > 0:
            fixed_cost_ratio = c_fixed / total_edge
        else:
            fixed_cost_ratio = 999

        # Minimum remaining volume check
        min_vol_pct = self.cfg.get('partial_exit_min_volume_pct', 0.30)
        min_vol_check = (v_rem / v_total) >= min_vol_pct

        # Edge check
        edge_check = ev_rem > 0

        # Fee burden check
        fee_burden_check = (cost_partial - cost_single) <= (total_edge * 0.10)

        # Fixed cost check
        fixed_cost_check = fixed_cost_ratio <= self.cfg.get('partial_exit_fixed_cost_ratio', 0.05)

        if self.debug:
            print(f"📊 PARTIAL EXIT EVALUATION:")
            print(f"   V_rem: {v_rem:.2f}, V_total: {v_total:.2f}")
            print(f"   EV_rem: ${ev_rem:.2f}")
            print(f"   Fixed Cost Ratio: {fixed_cost_ratio:.1%}")
            print(f"   Min Vol Check: {min_vol_check}")
            print(f"   Edge Check: {edge_check}")
            print(f"   Fee Burden: {fee_burden_check}")

        if not edge_check:
            return False, f"negative_ev_rem_{ev_rem:.2f}"
        if not min_vol_check:
            return False, f"min_vol_{v_rem/v_total:.1%}"
        if not fixed_cost_check:
            return False, f"fixed_cost_{fixed_cost_ratio:.1%}"
        if not fee_burden_check:
            return False, "fee_burden_too_high"

        return True, "partial_exit_allowed"

    def evaluate_exit(self, current_price, entry_price, stop_loss,
                      highest_price, lowest_price, bars_held, partial_exits,
                      ema_fast, ema_mid, ema_slow,
                      macd, macd_signal, macd_prev, signal_prev,
                      stoch_k, stoch_d, rsi, adx, atr,
                      position_type='long', trailing_activated=False,
                      trailing_stop=None, tier=2,
                      original_quantity=100, remaining_quantity=100):
        stop_key = (f"tier{1 if tier == 1 else 2}_stop_mult_"
                    f"{'short' if position_type == 'short' else 'long'}")
        default_mult = self.cfg.get('stop_loss_atr_mult', 2.2)
        sd = atr * self.cfg.get(stop_key, default_mult)
        pr = ((current_price - entry_price) / sd
              if position_type == 'long'
              else (entry_price - current_price) / sd) if sd > 0 else 0.
        mh = self.cfg.get('min_hold_bars_before_stop', 2)

        move_pct = ((current_price - entry_price) / entry_price
                    if position_type == 'long'
                    else (entry_price - current_price) / entry_price) if entry_price > 0 else 0.
        min_move_pct = (self.cfg.get('commission_round_trip_pct', 0.0010)
                        * self.cfg.get('early_exit_min_profit_multiple', 2.5))
        clears_fee_floor = move_pct >= min_move_pct

        # Hard stops
        if position_type == 'long':
            if current_price <= stop_loss and bars_held >= mh:
                return "stop_loss_hard", 1.
            if current_price <= stop_loss and bars_held < mh:
                if current_price <= entry_price - sd * self.cfg.get('emergency_stop_mult', 3.0):
                    return "stop_loss_emergency", 1.
        else:
            if current_price >= stop_loss and bars_held >= mh:
                return "stop_loss_hard", 1.
            if current_price >= stop_loss and bars_held < mh:
                if current_price >= entry_price + sd * self.cfg.get('emergency_stop_mult', 3.0):
                    return "stop_loss_emergency", 1.

        # Trailing stop
        if trailing_activated and trailing_stop is not None:
            if position_type == 'long' and current_price <= trailing_stop:
                return "trailing_stop", 1.
            if position_type == 'short' and current_price >= trailing_stop:
                return "trailing_stop", 1.

        # SMART PARTIAL EXIT WITH EVALUATION
        r1 = self.cfg.get('take_profit_r1', 1.5)
        r2 = self.cfg.get('take_profit_r2', 2.8)

        if partial_exits == 0 and pr >= r1:
            # Evaluate if partial exit makes financial sense
            partial_ok, partial_reason = self.evaluate_partial_exit(
                current_price, entry_price, position_type,
                original_quantity, remaining_quantity,
                partial_exits, pr, atr
            )

            if partial_ok:
                exit_pct = self.cfg.get('partial_exit_pct_r1', 0.50)
                return "partial_r1", exit_pct
            else:
                if self.debug:
                    print(f"📊 PARTIAL REJECTED: {partial_reason} (R={pr:.2f})")
                # If partial not allowed, check for full exit at R2
                if pr >= r2:
                    return "take_profit_full", 1.0
                # Otherwise continue holding

        if partial_exits == 1 and pr >= r2:
            # Evaluate second partial
            partial_ok, partial_reason = self.evaluate_partial_exit(
                current_price, entry_price, position_type,
                original_quantity, remaining_quantity,
                partial_exits, pr, atr
            )

            if partial_ok:
                exit_pct = self.cfg.get('partial_exit_pct_r2', 0.30)
                return "partial_r2", exit_pct
            else:
                # If second partial not profitable, take full exit
                return "take_profit_full", 1.0

        # Force full exit at R2
        if pr >= r2 and clears_fee_floor:
            return "take_profit_full", 1.0

        # Runner EMA reversal exit
        if (self.cfg.get('runner_ema_reversal_exit_enabled', True)):
            min_r = self.cfg.get('runner_ema_reversal_min_profit_r', 0.3)
            if pr >= min_r and clears_fee_floor:
                if position_type == 'long' and ema_fast < ema_mid < ema_slow:
                    return "runner_ema_reversal", 1.0
                if position_type == 'short' and ema_fast > ema_mid > ema_slow:
                    return "runner_ema_reversal", 1.0

        # RSI exits
        if position_type == 'long' and rsi > self.cfg.get('rsi_overbought_exit', 75) and pr >= .4 and clears_fee_floor:
            return "rsi_overbought", 1.
        if position_type == 'short' and rsi < self.cfg.get('rsi_oversold_exit', 25) and pr >= .4 and clears_fee_floor:
            return "rsi_oversold", 1.

        # Stoch reversal exits
        if (self.cfg.get('stoch_reversal_exit_enabled', True)
                and pr >= self.cfg.get('stoch_reversal_min_profit_r', .5)
                and clears_fee_floor):
            if position_type == 'long':
                ob = self.cfg.get('stoch_overbought', 80)
                if stoch_k > ob and stoch_d > ob and stoch_k < stoch_d:
                    return "stoch_overbought_cross", 1.
            else:
                os_ = self.cfg.get('stoch_oversold', 20)
                if stoch_k < os_ and stoch_d < os_ and stoch_k > stoch_d:
                    return "stoch_oversold_cross", 1.

        # MACD cross exits
        if (self.cfg.get('macd_cross_exit_enabled', True)
                and pr >= self.cfg.get('macd_cross_min_profit_r', .8)
                and clears_fee_floor):
            if position_type == 'long':
                if macd_prev >= signal_prev and macd < macd_signal:
                    if not (ema_fast > ema_slow and adx >= 28):
                        return "macd_bearish_cross", 1.
            else:
                if macd_prev <= signal_prev and macd > macd_signal:
                    if not (ema_fast < ema_slow and adx >= 28):
                        return "macd_bullish_cross", 1.

        # EMA cross exits
        if (self.cfg.get('ema_cross_exit_enabled', True)
                and pr >= self.cfg.get('ema_cross_min_profit_r', 1.5)
                and clears_fee_floor):
            if position_type == 'long' and ema_fast < ema_mid < ema_slow:
                return "ema_full_reversal", 1.
            if position_type == 'short' and ema_fast > ema_mid > ema_slow:
                return "ema_full_reversal", 1.

        # Max hold time
        if bars_held >= self.cfg.get('max_hold_bars', 24):
            return "max_hold_time", 1.

        return None, 0.


# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURATION MANAGER
# ═══════════════════════════════════════════════════════════════════════════

class ScalpingConfig:
    CONFIG_FILE = "scalping_settings.json"
    _custom_params = {}
    _current_mode = "Default Parameters"

    @classmethod
    def get_config(cls, override=None, timeframe=None, force_direction=None):
        config = SCALPING_PARAMS.copy()

        if os.path.exists(cls.CONFIG_FILE):
            try:
                with open(cls.CONFIG_FILE, 'r') as f:
                    saved = json.load(f)
                if saved.get('selected_mode') == "Custom Parameters":
                    for k, v in saved.get('custom_params', {}).items():
                        if k in config:
                            config[k] = v
            except Exception as e:
                logging.warning(f"Config load: {e}")

        if override and isinstance(override, dict):
            for k, v in override.items():
                if k in config:
                    config[k] = v

        if force_direction is not None:
            config['trade_direction'] = force_direction

        config['trade_direction'] = _normalize_trade_direction(config.get('trade_direction', 'both'))
        DIRECTION_MANAGER.set_direction(config['trade_direction'])
        logging.info(f"📊 ScalpingConfig: Direction = {config['trade_direction'].upper()}")

        if timeframe is None:
            timeframe = GlobalConfig.ACTIVE_TIMEFRAME
        config = scale_params_for_timeframe(config, timeframe)
        config['timeframe'] = timeframe
        config['bar_interval_minutes'] = GlobalConfig.TIMEFRAME_MINUTES.get(timeframe, 60)

        return config

    @classmethod
    def update_gui_direction(cls, direction: str):
        DIRECTION_MANAGER.set_direction(direction)
        logging.info(f"📊 ScalpingConfig: GUI direction updated to {direction.upper()}")

    @classmethod
    def save_config(cls, custom_params, selected_mode="Custom Parameters"):
        try:
            with open(cls.CONFIG_FILE, 'w') as f:
                json.dump({'timestamp': datetime.now().isoformat(),
                           'selected_mode': selected_mode,
                           'custom_params': custom_params}, f, indent=4)
            return True
        except Exception as e:
            logging.error(f"Config save: {e}")
            return False

    @classmethod
    def reset_to_defaults(cls):
        return SCALPING_PARAMS.copy()


# ═══════════════════════════════════════════════════════════════════════════
# INDICATOR CALCULATOR
# ═══════════════════════════════════════════════════════════════════════════

class IndicatorCalculator:

    @staticmethod
    def calculate(df: pd.DataFrame, params: dict) -> pd.DataFrame:
        df = df.copy()
        try:
            df['EMA_Fast'] = talib.EMA(df['Close'], params['ema_fast_period'])
            df['EMA_Mid'] = talib.EMA(df['Close'], params['ema_mid_period'])
            df['EMA_Slow'] = talib.EMA(df['Close'], params['ema_slow_period'])
            df['EMA_Daily'] = talib.EMA(df['Close'], params.get('daily_ema_period', 24))
            df['Above_Daily'] = (df['Close'] > df['EMA_Daily']).astype(bool)

            df['MACD'], df['MACD_Signal'], df['MACD_Histogram'] = talib.MACD(
                df['Close'],
                fastperiod=params['macd_fast'],
                slowperiod=params['macd_slow'],
                signalperiod=params['macd_signal_period'])
            df['MACD_Hist_Rising'] = df['MACD_Histogram'] > df['MACD_Histogram'].shift(1)

            df['Stoch_K'], df['Stoch_D'] = talib.STOCH(
                df['High'], df['Low'], df['Close'],
                fastk_period=params['stoch_k_period'],
                slowk_period=params['stoch_smooth'], slowk_matype=0,
                slowd_period=params['stoch_d_period'], slowd_matype=0)
            df['Stoch_K_Rising'] = df['Stoch_K'] > df['Stoch_K'].shift(1)

            df['ADX'] = talib.ADX(df['High'], df['Low'], df['Close'], params['adx_period'])
            df['ADX_prev'] = df['ADX'].shift(1)
            df['DMP'] = talib.PLUS_DI(df['High'], df['Low'], df['Close'], params['adx_period'])
            df['DMM'] = talib.MINUS_DI(df['High'], df['Low'], df['Close'], params['adx_period'])
            df['RSI'] = talib.RSI(df['Close'], params['rsi_period'])

            df['Volume_MA'] = talib.SMA(df['Volume'], params['volume_period'])
            with np.errstate(divide='ignore', invalid='ignore'):
                df['Volume_Ratio'] = np.where(df['Volume_MA'] > 0,
                                              df['Volume'] / df['Volume_MA'], 1.)
            df['Volume_Ratio'] = (df['Volume_Ratio']
                                  .replace([np.inf, -np.inf], np.nan)
                                  .fillna(1.).clip(.01, 10.))

            df['ATR'] = talib.ATR(df['High'], df['Low'], df['Close'], params['atr_period'])
            df['ATR_MA'] = df['ATR'].rolling(params.get('atr_compression_lookback', 40)).mean()
            df['ATR_Compressed'] = (
                    df['ATR'] < df['ATR_MA'] * params.get('atr_compression_threshold', .35))
            spike_ratio = params.get('atr_spike_ratio', 1.8)
            df['ATR_Spike'] = (df['ATR'] > df['ATR_MA'] * spike_ratio).fillna(False)

            df['Momentum'] = df['Close'].pct_change(params.get('momentum_period', 3)) * 100

            df = IndicatorCalculator._detect_ranging(df, params)

            lb = max(20, min(50, params.get('extended_run_lookback', 6) * 2))
            df['High_20'] = df['High'].rolling(lb).max()
            df['Low_20'] = df['Low'].rolling(lb).min()
            df['Price_Range_20'] = df['High_20'] - df['Low_20']
            with np.errstate(divide='ignore', invalid='ignore'):
                df['Price_Pct_20'] = np.where(
                    df['Price_Range_20'] > 0,
                    (df['Close'] - df['Low_20']) / df['Price_Range_20'] * 100, 50.)
            df['Price_Pct_20'] = df['Price_Pct_20'].clip(0, 100).fillna(50)

            rl = params.get('extended_run_lookback', 6)
            df['Swing_Low'] = df['Low'].rolling(rl).min()
            df['Swing_High'] = df['High'].rolling(rl).max()
            df['Run_From_Low'] = ((df['Close'] - df['Swing_Low']) / df['Swing_Low'] * 100).fillna(0)
            df['Run_From_High'] = ((df['Swing_High'] - df['Close']) / df['Swing_High'] * 100).fillna(0)

            bull = (df['EMA_Fast'] > df['EMA_Slow']).astype(int)
            bear = (df['EMA_Fast'] < df['EMA_Slow']).astype(int)
            df['Trend_Age_Bull'] = bull.groupby(bull.ne(bull.shift()).cumsum()).cumsum()
            df['Trend_Age_Bear'] = bear.groupby(bear.ne(bear.shift()).cumsum()).cumsum()

            fb = ((df['EMA_Fast'] > df['EMA_Mid']) & (df['EMA_Mid'] > df['EMA_Slow'])).astype(int)
            bb2 = ((df['EMA_Fast'] < df['EMA_Mid']) & (df['EMA_Mid'] < df['EMA_Slow'])).astype(int)
            df['Full_Bull_Age'] = fb.groupby(fb.ne(fb.shift()).cumsum()).cumsum()
            df['Full_Bear_Age'] = bb2.groupby(bb2.ne(bb2.shift()).cumsum()).cumsum()

            for col in ['EMA_Fast', 'EMA_Mid', 'EMA_Slow', 'ADX', 'RSI', 'MACD', 'MACD_Signal',
                        'MACD_Histogram', 'Stoch_K', 'Stoch_D', 'Volume_Ratio', 'ATR', 'Momentum',
                        'Price_Pct_20', 'Above_Daily', 'ATR_Compressed', 'ATR_Spike', 'Ranging',
                        'Trend_Age_Bull', 'Trend_Age_Bear', 'Full_Bull_Age', 'Full_Bear_Age']:
                if col in df.columns:
                    df[f'{col}_closed'] = df[col].shift(1)
            return df
        except Exception as e:
            logging.error(f"Indicator error: {e}")
            raise

    @staticmethod
    def _detect_ranging(df, params):
        df = df.copy()
        bb_p = params.get('bb_period', 20)
        bb_s = params.get('bb_std', 2.)
        df['BB_Upper'], df['BB_Mid'], df['BB_Lower'] = talib.BBANDS(
            df['Close'], timeperiod=bb_p, nbdevup=bb_s, nbdevdn=bb_s, matype=0)
        df['BB_Width'] = (df['BB_Upper'] - df['BB_Lower']) / df['Close'].replace(0, 1)
        kp = params.get('kc_period', 20)
        km = params.get('kc_atr_mult', 1.5)
        df['KC_Mid'] = df['Close'].ewm(span=kp).mean()
        df['KC_ATR'] = (df['ATR'].rolling(kp).mean()
                        if 'ATR' in df.columns
                        else (df['BB_Upper'] - df['BB_Lower']) / 4)
        df['KC_Upper'] = df['KC_Mid'] + km * df['KC_ATR']
        df['KC_Lower'] = df['KC_Mid'] - km * df['KC_ATR']
        df['KC_Width'] = df['KC_Upper'] - df['KC_Lower']
        df['CHOP'] = IndicatorCalculator._choppiness(
            df['High'], df['Low'], df['Close'], params.get('chop_period', 14))
        mc = params.get('ranging_min_checks', 4)
        c1 = (abs(df['Close'] - df['EMA_Fast']) / df['EMA_Fast'].replace(0, 1) <= .004).fillna(False)
        c2 = (df['BB_Width'] < df['KC_Width']).fillna(False)
        c3 = (df['ADX'] < 20).fillna(False) if 'ADX' in df.columns else pd.Series(False, index=df.index)
        c4 = df['RSI'].between(44, 56).fillna(False) if 'RSI' in df.columns else pd.Series(False, index=df.index)
        c5 = (df['CHOP'] >= params.get('chop_threshold', 61)).fillna(False)
        c6 = (df['Volume_Ratio'] < 0.8).fillna(False) if 'Volume_Ratio' in df.columns else pd.Series(False,
                                                                                                     index=df.index)
        df['Ranging'] = (c1.astype(int) + c2.astype(int) + c3.astype(int) +
                         c4.astype(int) + c5.astype(int) + c6.astype(int) >= mc).fillna(False)
        return df

    @staticmethod
    def _choppiness(h, l, c, p=14):
        tr = pd.concat([h - l, abs(h - c.shift(1)), abs(l - c.shift(1))], axis=1).max(axis=1)
        return (100 * np.log10(tr.rolling(p).sum() /
                               (h.rolling(p).max() - l.rolling(p).min()).replace(0, 1))
                / np.log10(p)).fillna(50)


# ═══════════════════════════════════════════════════════════════════════════
# CORE SCALPING LOGIC
# ═══════════════════════════════════════════════════════════════════════════

class ScalpingLogic:

    def __init__(self, config=None, trading_app=None):
        self.config = config or ScalpingConfig.get_config()
        self.trading_app = trading_app
        self.timeframe = (trading_app.get_timeframe()
                          if trading_app and hasattr(trading_app, 'get_timeframe')
                          else self.config.get('timeframe', '1h'))
        if self.config.get('timeframe') != self.timeframe:
            self.config = ScalpingConfig.get_config(self.config, self.timeframe)

        if trading_app and hasattr(trading_app, 'trade_direction_var'):
            try:
                self.trade_direction = _normalize_trade_direction(trading_app.trade_direction_var.get())
            except Exception as e:
                logging.warning(f"Could not read trading_app direction: {e}")
                self.trade_direction = _normalize_trade_direction(self.config.get('trade_direction', 'both'))
        else:
            self.trade_direction = _normalize_trade_direction(self.config.get('trade_direction', 'both'))

        DIRECTION_MANAGER.set_direction(self.trade_direction)

        self.only_long_entries = (self.trade_direction == 'long')
        self.only_short_entries = (self.trade_direction == 'short')

        self.config['trade_direction'] = self.trade_direction

        for k, v in self.config.items():
            setattr(self, k, v)

        self._quality_min_long_default = self.quality_min_long
        self._quality_min_short_default = self.quality_min_short
        self.risk_controller = ScalpingRiskController()
        self.risk_controller.max_position_units = getattr(self, 'max_position_units', 100)
        self.regime_detector = RegimeDetector()
        self.exit_manager = ScalpingExitManager(self.config)
        self.strategy_state = StrategyState.SEEKING_ENTRY
        self.bar_count = 0
        self.bars_held = 0
        self.last_trade_bar = -999
        self.trade_counter = 0
        self.total_trades = 0
        self.winning_trades = 0
        self.losing_trades = 0
        self.total_profit = 0.
        self.trade_history: List[dict] = []
        self.current_regime = "UNKNOWN"
        self.equity_curve = [GlobalConfig.INITIAL_CAPITAL]
        self._current_df = None
        self._pending_signal = None
        self._signal_bar = -999
        self._signal_price = None
        self._last_profit_target_bar = -999
        self._consecutive_loss_count = 0
        self._last_loss_bar = -999
        self._last_quality_score = 0
        self._near_miss_trades = []
        self._last_dir_cache = {'bar': -999, 'dir': None, 'result': True, 'reason': ''}

        self._maker_order_pending = False
        self._maker_order_timeout = -999
        self._maker_order_price = 0.0
        self._maker_order_direction = ''

        self._log(f"📊 v1.7.3 {self.timeframe} | Dir:{self.trade_direction.upper()} "
                  f"| T1 Q≥{getattr(self, 'quality_tier1_min_long', 80)}(L)/"
                  f"{getattr(self, 'quality_tier1_min_short', 80)}(S) "
                  f"| T2 Q≥{getattr(self, 'quality_tier2_min_long', 72)}(L)/"
                  f"{getattr(self, 'quality_tier2_min_short', 72)}(S) "
                  f"| Age≥{getattr(self, 'trend_age_min_bars', 3)}b "
                  f"| Maker Orders: {'ON' if getattr(self, 'use_maker_orders', True) else 'OFF'}", "cyan")

    def _to_in_trade(self):
        self.strategy_state = StrategyState.IN_TRADE

    def _to_seeking_entry(self):
        self.strategy_state = StrategyState.SEEKING_ENTRY

    def _transition_to_in_trade(self):
        self._to_in_trade()

    def _transition_to_seeking_entry(self):
        self._to_seeking_entry()

    def _log(self, msg, color="white"):
        if self.trading_app and hasattr(self.trading_app, 'log_message'):
            self.trading_app.log_message(msg, color)
        else:
            print(f"[{color.upper()}] {msg}")

    def _tier_stop_mult(self, tier: int, direction: str) -> float:
        if tier == 1:
            return (getattr(self, 'tier1_stop_mult_short', 2.4) if direction == 'short'
                    else getattr(self, 'tier1_stop_mult_long', 1.8))
        return (getattr(self, 'tier2_stop_mult_short', 2.8) if direction == 'short'
                else getattr(self, 'tier2_stop_mult_long', 2.2))

    def _tier_trailing_params(self, tier: int) -> Tuple[float, float]:
        if tier == 1:
            return (getattr(self, 'trailing_activation_tier1', 0.016),
                    getattr(self, 'trailing_distance_tier1', 0.010))
        return (getattr(self, 'trailing_activation_tier2', 0.024),
                getattr(self, 'trailing_distance_tier2', 0.014))

    def _is_atr_compressed(self, data):
        df = self._current_df
        if df is None or len(df) < 40: return False
        an = data.get('ATR', 0)
        aa = float(df['ATR_MA'].iloc[-1]) if 'ATR_MA' in df.columns else float(df['ATR'].iloc[-40:].mean())
        if aa <= 0: return False
        c = an < aa * getattr(self, 'atr_compression_threshold', .35)
        if c:
            r = df.iloc[-1]
            if (float(r.get('EMA_Fast', 0)) > float(r.get('EMA_Mid', 0)) > float(r.get('EMA_Slow', 0))
                    and float(r.get('ADX', 0)) >= 25):
                return False
        return c

    def _current_atr_regime_args(self, data) -> dict:
        df = self._current_df
        atr_now = data.get('ATR', 0.)
        atr_avg = 0.
        if df is not None and 'ATR_MA' in df.columns:
            try:
                atr_avg = float(df['ATR_MA'].iloc[-1])
            except:
                pass
        elif df is not None and 'ATR' in df.columns and len(df) >= 20:
            atr_avg = float(df['ATR'].iloc[-20:].mean())
        return {'atr_now': atr_now, 'atr_avg': atr_avg}

    def _is_extended_run_long(self, data):
        df = self._current_df
        if df is None or len(df) < 12: return False
        c = data.get('Close', 0)
        sl = float(df['Swing_Low'].iloc[-1]) if 'Swing_Low' in df.columns else float(df['Low'].iloc[-12:].min())
        if sl <= 0: return False
        if (c - sl) / sl * 100 > getattr(self, 'extended_run_max_pct_long', 5.):
            self._log("🔴 EXT_RUN_long", "red")
            return True
        return False

    def _is_extended_run_short(self, data):
        df = self._current_df
        if df is None or len(df) < 12: return False
        c = data.get('Close', 0)
        sh = float(df['Swing_High'].iloc[-1]) if 'Swing_High' in df.columns else float(df['High'].iloc[-12:].max())
        if sh <= 0: return False
        if (sh - c) / sh * 100 > getattr(self, 'extended_run_max_pct_short', 5.):
            self._log("🔴 EXT_RUN_short", "red")
            return True
        return False

    def _is_consecutive_loss_cooldown(self):
        thresh = getattr(self, 'consecutive_loss_threshold', 3)
        cd = getattr(self, 'consecutive_loss_cooldown_bars', 5)
        if self._consecutive_loss_count >= thresh:
            if (self.bar_count - self._last_loss_bar) < cd:
                return True
            else:
                self._consecutive_loss_count = 0
        return False

    def _daily_trend_is_up(self, data):
        if not getattr(self, 'daily_trend_filter_enabled', True): return True
        above = data.get('Above_Daily', None)
        if above is None:
            df = self._current_df
            above = bool(df['Above_Daily'].iloc[-1]) if df is not None and 'Above_Daily' in df.columns else True
        if bool(above): return True
        ef, em, es, adx = (data.get('EMA_Fast', 0), data.get('EMA_Mid', 0),
                           data.get('EMA_Slow', 0), data.get('ADX', 0))
        return ef > em > es and adx >= getattr(self, 'daily_trend_adx_override', 20)

    def _daily_trend_is_down(self, data):
        if not getattr(self, 'daily_trend_filter_enabled', True): return True
        above = data.get('Above_Daily', None)
        if above is None:
            df = self._current_df
            above = bool(df['Above_Daily'].iloc[-1]) if df is not None and 'Above_Daily' in df.columns else False
        return not bool(above)

    def _get_full_stack_age(self, direction):
        df = self._current_df
        if df is None or len(df) < 2: return 0
        col = 'Full_Bull_Age' if direction == 'long' else 'Full_Bear_Age'
        if col in df.columns: return int(df[col].iloc[-1])
        count = 0
        for i in range(len(df) - 1, max(len(df) - 50, -1), -1):
            r = df.iloc[i]
            ef, em, es = r.get('EMA_Fast', 0), r.get('EMA_Mid', 0), r.get('EMA_Slow', 0)
            ok = (ef > em > es) if direction == 'long' else (ef < em < es)
            if ok:
                count += 1
            else:
                break
        return count

    def _quality_score_long(self, data):
        s = 0
        comp = {}
        bk = []
        ef, em, es = data.get('EMA_Fast', 0), data.get('EMA_Mid', 0), data.get('EMA_Slow', 0)
        if ef > em > es:
            s += 35
            comp['ema'] = 35
            bk.append("EMA_FULL+35")
        elif ef > es:
            s += 15
            comp['ema'] = 15
            bk.append("EMA_PART+15")
        else:
            bk.append("EMA_BEAR+0")
        price = data.get('Close', 0)
        if price > ef:
            s += 10
            comp['p_fast'] = 10
            bk.append("P>EF+10")
        if price > em:
            s += 5
            comp['p_mid'] = 5
            bk.append("P>EM+5")
        macd, ms, mh = data.get('MACD', 0), data.get('MACD_Signal', 0), data.get('MACD_Histogram', 0)
        if macd > ms:
            s += 15
            comp['macd'] = 15
            bk.append("MACD>SIG+15")
        if mh > 0:
            s += 5
            comp['mh'] = 5
            bk.append("HIST+5")
        rsi = data.get('RSI', 50)
        if getattr(self, 'rsi_long_min', 40) <= rsi <= getattr(self, 'rsi_long_max', 70):
            s += 15
            comp['rsi'] = 15
            bk.append(f"RSI+15({rsi:.0f})")
        vr = data.get('Volume_Ratio', 1.)
        if vr >= getattr(self, 'volume_strong_ratio', 1.6):
            s += 10
            comp['vol'] = 10
            bk.append("VOL_STR+10")
        elif vr >= getattr(self, 'volume_min_ratio', .9):
            s += 5
            comp['vol'] = 5
            bk.append("VOL_OK+5")
        adx, dmp, dmm = data.get('ADX', 20), data.get('DMP', 0), data.get('DMM', 0)
        if adx >= getattr(self, 'adx_min_long', 18):
            s += 5
            comp['adx'] = 5
            bk.append(f"ADX+5({adx:.0f})")
            if dmp > dmm:
                s += 5
                comp['di'] = 5
                bk.append("+DI>-DI+5")
        total = min(s, 100)
        self._log(f"long Q={total} — {' | '.join(bk)}", "cyan")
        return total, comp, " | ".join(bk)

    def _quality_score_short(self, data):
        s = 0
        comp = {}
        bk = []
        ef, em, es = data.get('EMA_Fast', 0), data.get('EMA_Mid', 0), data.get('EMA_Slow', 0)
        if ef < em < es:
            s += 35
            comp['ema'] = 35
            bk.append("EMA_FULL+35")
        elif ef < es:
            s += 15
            comp['ema'] = 15
            bk.append("EMA_PART+15")
        else:
            bk.append("EMA_BULL+0")
        price = data.get('Close', 0)
        if price < ef:
            s += 10
            comp['p_fast'] = 10
            bk.append("P<EF+10")
        if price < em:
            s += 5
            comp['p_mid'] = 5
            bk.append("P<EM+5")
        macd, ms, mh = data.get('MACD', 0), data.get('MACD_Signal', 0), data.get('MACD_Histogram', 0)
        if macd < ms:
            s += 15
            comp['macd'] = 15
            bk.append("MACD<SIG+15")
        if mh < 0:
            s += 5
            comp['mh'] = 5
            bk.append("HIST_NEG+5")
        rsi = data.get('RSI', 50)
        if getattr(self, 'rsi_short_min', 30) <= rsi <= getattr(self, 'rsi_short_max', 60):
            s += 15
            comp['rsi'] = 15
            bk.append(f"RSI+15({rsi:.0f})")
        vr = data.get('Volume_Ratio', 1.)
        if vr >= getattr(self, 'volume_strong_ratio', 1.6):
            s += 10
            comp['vol'] = 10
            bk.append("VOL_STR+10")
        elif vr >= getattr(self, 'volume_min_ratio', .9):
            s += 5
            comp['vol'] = 5
            bk.append("VOL_OK+5")
        adx, dmp, dmm = data.get('ADX', 20), data.get('DMP', 0), data.get('DMM', 0)
        if adx >= getattr(self, 'adx_min_short', 18):
            s += 5
            comp['adx'] = 5
            bk.append(f"ADX+5({adx:.0f})")
            if dmm > dmp:
                s += 5
                comp['di'] = 5
                bk.append("-DI>+DI+5")
        total = min(s, 100)
        self._log(f"short Q={total} — {' | '.join(bk)}", "cyan")
        return total, comp, " | ".join(bk)

    def _check_long_filters(self, data):
        ef, em, es = data.get('EMA_Fast', 0), data.get('EMA_Mid', 0), data.get('EMA_Slow', 0)
        if not (ef > em > es): return "long_no_full_stack"
        ma = getattr(self, 'trend_age_min_bars', 3)
        if self._get_full_stack_age('long') < ma: return "long_trend_young"
        adx = data.get('ADX', 0)
        if adx < getattr(self, 'adx_min_long', 18): return f"long_adx_weak_{adx:.1f}"
        if adx > getattr(self, 'adx_extended_threshold', 55): return f"long_adx_ext_{adx:.1f}"
        rsi = data.get('RSI', 50)
        rl = getattr(self, 'rsi_long_min', 40)
        rh = getattr(self, 'rsi_long_max', 70)
        if not (rl <= rsi <= rh): return f"long_rsi_{rsi:.0f}"
        vr = float(data.get('Volume_Ratio', 1.))
        if vr < getattr(self, 'volume_min_ratio', .9): return f"long_vol_{vr:.2f}"
        if data.get('MACD', 0) <= data.get('MACD_Signal', 0): return "long_macd_below_sig"
        sk, sd = data.get('Stoch_K', 50), data.get('Stoch_D', 50)
        if sk <= sd: return "long_sk_lte_sd"
        if sk > 75: return f"long_sk_too_high_{sk:.0f}"
        if sk > getattr(self, 'stoch_overbought', 80) and sd > getattr(self, 'stoch_overbought', 80):
            return "long_stoch_ob"
        if data.get('Momentum', 0) < getattr(self, 'momentum_min_long', .003):
            return "long_mom_weak"
        if ef > 0:
            dist = (data.get('Close', 0) - ef) / ef * 100
            pzl = getattr(self, 'pullback_zone_lower_pct', -3.)
            pzu = getattr(self, 'pullback_zone_upper_pct', 1.5)
            if not (pzl <= dist <= pzu):
                return f"long_zone_{dist:.2f}pct"
        adx_now = data.get('ADX', 0)
        adx_prev = data.get('ADX_prev', adx_now)
        bm = getattr(self, 'bar_interval_minutes', None) or self.config.get('bar_interval_minutes', None) or 60
        asl = getattr(self, 'adx_slope_min', -.5)
        if bm <= 15: asl = -999
        if (adx_now - adx_prev) < asl: return "long_adx_slope_falling"
        if not self._daily_trend_is_up(data): return "long_daily_trend_down"
        return "pass"

    def _check_short_filters(self, data):
        ef, em, es = data.get('EMA_Fast', 0), data.get('EMA_Mid', 0), data.get('EMA_Slow', 0)
        if not (ef < em < es): return "short_no_full_stack"
        ma = getattr(self, 'trend_age_min_bars', 3)
        if self._get_full_stack_age('short') < ma: return "short_trend_young"
        adx = data.get('ADX', 0)
        if adx < getattr(self, 'adx_min_short', 18): return f"short_adx_weak_{adx:.1f}"
        if adx > getattr(self, 'adx_extended_threshold', 55): return f"short_adx_ext_{adx:.1f}"
        rsi = data.get('RSI', 50)
        rl = getattr(self, 'rsi_short_min', 30)
        rh = getattr(self, 'rsi_short_max', 60)
        if not (rl <= rsi <= rh): return f"short_rsi_{rsi:.0f}"
        vr = float(data.get('Volume_Ratio', 1.))
        if vr < getattr(self, 'volume_min_ratio', .9): return f"short_vol_{vr:.2f}"
        if data.get('MACD', 0) >= data.get('MACD_Signal', 0): return "short_macd_above_sig"
        sk, sd = data.get('Stoch_K', 50), data.get('Stoch_D', 50)
        if sk >= sd: return "short_sk_gte_sd"
        if sk < 25: return f"short_sk_too_low_{sk:.0f}"
        if sk < getattr(self, 'stoch_oversold', 20) and sd < getattr(self, 'stoch_oversold', 20):
            return "short_stoch_os"
        if data.get('Momentum', 0) > -getattr(self, 'momentum_min_short', .003):
            return "short_mom_not_bear"
        if ef > 0:
            dist = (data.get('Close', 0) - ef) / ef * 100
            lower = getattr(self, 'pullback_zone_lower_pct', -3.)
            upper = getattr(self, 'pullback_zone_upper_pct', 1.5)
            if not (lower <= dist <= upper):
                return f"short_zone_{dist:.2f}pct"
        adx_now = data.get('ADX', 0)
        adx_prev = data.get('ADX_prev', adx_now)
        bm = getattr(self, 'bar_interval_minutes', None) or self.config.get('bar_interval_minutes', None) or 60
        asl = getattr(self, 'adx_slope_min', -.5)
        if bm <= 15: asl = -999
        if (adx_now - adx_prev) < asl: return "short_adx_slope_falling"
        if not self._daily_trend_is_down(data): return "short_daily_trend_up"
        return "pass"

    def _create_pending_signal(self, direction, quality, tier, component_scores, breakdown, data):
        decision = 'buy' if direction == 'long' else 'sell_short'
        self._pending_signal = {
            'direction': direction, 'decision': decision,
            'quality_score': quality, 'tier': tier,
            'position_mult': 1.,
            'breakdown': breakdown, 'component_scores': component_scores,
            'signal_price': data.get('Close', 0), 'signal_adx': data.get('ADX', 0),
            'signal_rsi': data.get('RSI', 50), 'signal_macd': data.get('MACD', 0),
            'signal_volume': data.get('Volume_Ratio', 1.),
            'signal_price_pct': data.get('Price_Pct_20', 50),
            'signal_bar': self.bar_count,
            'signal_time': datetime.now(timezone.utc),
        }
        self._signal_bar = self.bar_count
        self._signal_price = data.get('Close', 0)
        icon = "📈" if direction == 'long' else "📉"
        self._log(f"{icon} T{tier} {direction.upper()} SIGNAL @${self._signal_price:.4f} "
                  f"Q={quality}→next bar", "purple")
        return ("hold", quality, f"PENDING:{direction}:{tier}", component_scores)

    def _daily_trade_limit_reached(self):
        limit = getattr(self, 'max_daily_trades', None)
        if not limit or limit <= 0:
            return False
        try:
            current_date = self.data.index[-1].date()
        except Exception:
            return False
        if getattr(self, '_daily_trade_date', None) != current_date:
            self._daily_trade_date = current_date
            self._daily_trade_count = 0
        return self._daily_trade_count >= limit

    def _register_daily_trade(self):
        try:
            current_date = self.data.index[-1].date()
        except Exception:
            return
        if getattr(self, '_daily_trade_date', None) != current_date:
            self._daily_trade_date = current_date
            self._daily_trade_count = 0
        self._daily_trade_count = getattr(self, '_daily_trade_count', 0) + 1

    def update_trade_direction(self, new_direction: str):
        norm = _normalize_trade_direction(new_direction, default=None)
        if norm is None:
            self._log(f"⚠️ Invalid trade direction: {new_direction}", "orange")
            return
        new_direction = norm

        old_direction = self.trade_direction

        DIRECTION_MANAGER.set_direction(new_direction)

        self.trade_direction = new_direction
        self.only_long_entries = (new_direction == 'long')
        self.only_short_entries = (new_direction == 'short')

        if hasattr(self, 'config'):
            self.config['trade_direction'] = new_direction

        if self._pending_signal and new_direction != 'both':
            sig_dir = self._pending_signal.get('direction', '')
            if sig_dir and sig_dir != new_direction:
                self._pending_signal = None
                self._signal_bar = -999
                self._log(f"🔄 Cleared pending {sig_dir} signal (direction changed to {new_direction})", "orange")

        self._last_dir_cache = {'bar': -999, 'dir': None, 'result': True, 'reason': ''}

        self._log(f"📊 Direction updated: {old_direction} → {new_direction.upper()}", "bold cyan")

    def _should_trade_direction(self, direction: str) -> bool:
        current_direction = getattr(self, 'trade_direction', 'both')

        if current_direction == 'both':
            return True
        if current_direction == 'long':
            return direction == 'long'
        if current_direction == 'short':
            return direction == 'short'

        self._log(f"⚠️ Unknown trade_direction: {current_direction}, defaulting to allowing", "orange")
        return True

    def _check_entry_conditions(self, data):
        mb = getattr(self, 'min_bars_between_trades', 2)
        if (self.bar_count - self.last_trade_bar) < mb:
            return "hold", 0, "min_bar_gap", {}
        if self._daily_trade_limit_reached():
            return "hold", 0, "max_daily_trades", {}
        if self._is_atr_compressed(data):
            return "hold", 0, "atr_compressed", {}
        if self._is_consecutive_loss_cooldown():
            return "hold", 0, "loss_cooldown", {}
        if data.get('Ranging', False) and getattr(self, 'regime_filter_enabled', True):
            self._log("📊 RANGING — blocked", "orange")
            return "hold", 0, "ranging", {}
        if data.get('ATR_Spike', False) and getattr(self, 'regime_filter_enabled', True):
            self._log("⚡ ATR SPIKE — blocked", "orange")
            return "hold", 0, "atr_spike", {}

        my_dir = getattr(self, 'trade_direction', 'both')

        if my_dir == 'long':
            return self._entry_long(data)
        elif my_dir == 'short':
            return self._entry_short(data)
        else:
            return self._entry_both(data)

    def _entry_long(self, data):
        if self._is_extended_run_long(data): return "hold", 0, "ext_run_long", {}
        quality, comp, bkd = self._quality_score_long(data)
        self._last_quality_score = quality
        qm = getattr(self, 'quality_tier2_min_long', getattr(self, 'quality_min_long', 72))
        if quality < qm: return "hold", quality, f"Q_low_{quality}<{qm}", comp
        f = self._check_long_filters(data)
        if f != "pass": return "hold", quality, f"filter:{f}", comp
        atr_args = self._current_atr_regime_args(data)
        regime, _ = self.regime_detector.detect_regime(
            data.get('EMA_Fast', 0), data.get('EMA_Slow', 0), data.get('ADX', 0),
            **atr_args)
        self.current_regime = regime
        if not self.regime_detector.is_tradeable(regime, 'long'):
            return "hold", quality, f"regime_{regime}", comp
        t1 = getattr(self, 'quality_tier1_min_long', 80)
        tier = 1 if quality >= t1 else 2
        return self._create_pending_signal('long', quality, tier, comp, bkd, data)

    def _entry_short(self, data):
        if self._is_extended_run_short(data): return "hold", 0, "ext_run_short", {}
        quality, comp, bkd = self._quality_score_short(data)
        self._last_quality_score = quality
        qm = getattr(self, 'quality_tier2_min_short', getattr(self, 'quality_min_short', 72))
        if quality < qm: return "hold", quality, f"Q_low_{quality}<{qm}", comp
        f = self._check_short_filters(data)
        if f != "pass": return "hold", quality, f"filter:{f}", comp
        atr_args = self._current_atr_regime_args(data)
        regime, _ = self.regime_detector.detect_regime(
            data.get('EMA_Fast', 0), data.get('EMA_Slow', 0), data.get('ADX', 0),
            **atr_args)
        self.current_regime = regime
        if not self.regime_detector.is_tradeable(regime, 'short'):
            return "hold", quality, f"regime_{regime}", comp
        t1 = getattr(self, 'quality_tier1_min_short', 80)
        tier = 1 if quality >= t1 else 2
        return self._create_pending_signal('short', quality, tier, comp, bkd, data)

    def _entry_both(self, data):
        sp = self._pending_signal
        sb_ = self._signal_bar
        lr = self._entry_long(data)
        lp = self._pending_signal
        lb_ = self._signal_bar
        self._pending_signal = sp
        self._signal_bar = sb_
        sr = self._entry_short(data)
        short_p = self._pending_signal
        short_b = self._signal_bar

        la, lq, ll, lc = lr
        sa, sq, sl_, sc = sr

        lpp = bool(ll and ll.startswith("PENDING:"))
        spp = bool(sl_ and sl_.startswith("PENDING:"))
        lq = lq if isinstance(lq, (int, float)) else 0
        sq = sq if isinstance(sq, (int, float)) else 0

        if lpp and spp:
            if lq >= sq:
                self._pending_signal = lp
                self._signal_bar = lb_
                self._log(f"⚖️ long Q{lq}>Q{sq}", "cyan")
                return lr
            else:
                self._pending_signal = short_p
                self._signal_bar = short_b
                self._log(f"⚖️ short Q{sq}>Q{lq}", "cyan")
                return sr

        if lpp:
            self._pending_signal = lp
            self._signal_bar = lb_
            return lr
        if spp:
            self._pending_signal = short_p
            self._signal_bar = short_b
            return sr

        self._pending_signal = sp
        self._signal_bar = sb_

        if lq >= sq and lq > 0:
            reason = ll if ll else "no_clear_dir"
            return "hold", lq, reason, lc
        elif sq > 0:
            reason = sl_ if sl_ else "no_clear_dir"
            return "hold", sq, reason, sc
        return "hold", 0, "no_clear_dir", {}

    def calculate_position_size(self, equity, atr, price, quality_score=70, tier=1,
                                position_mult=1., direction=None):
        if direction is None:
            direction = (self._pending_signal.get('direction', 'long')
                         if self._pending_signal else 'long')
        stop_mult = self._tier_stop_mult(tier, direction)
        sd = atr * stop_mult
        sp = price - sd if direction == 'long' else price + sd
        adx = 25.
        if self._current_df is not None:
            try:
                adx = float(self._current_df['ADX'].iloc[-1])
            except:
                pass
        base = self.risk_controller.calculate_position_size(
            price, sp, quality_score, adx, tier=tier, direction=direction)
        rm = self.regime_detector.get_position_multiplier(self.current_regime)
        size = min(max((equity * .001) / price if price > 0 else 0,
                       base * rm * position_mult),
                   getattr(self, 'max_position_units', 100))
        if price >= 1000:
            return round(size, 6)
        elif price >= 100:
            return round(size, 4)
        else:
            return max(0, int(size))

    def record_trade(self, profit=0., exit_reason="unknown", tier=None, size=None,
                     direction=None, entry_quality=None, entry_price=None,
                     exit_price=None, hold_duration=None,
                     entry_bar=None, exit_bar=None, **kwargs):
        self.total_trades += 1
        self.total_profit += profit
        if profit > 0:
            self.winning_trades += 1
            self._consecutive_loss_count = 0
        else:
            self.losing_trades += 1
            self._consecutive_loss_count += 1
            self._last_loss_bar = self.bar_count
        self.last_trade_bar = self.bar_count
        self.trade_history.append({
            'profit': profit, 'exit_reason': exit_reason, 'tier': tier, 'size': size,
            'direction': direction, 'entry_quality': entry_quality, 'entry_price': entry_price,
            'exit_price': exit_price, 'hold_duration': hold_duration,
            'entry_bar': entry_bar, 'exit_bar': exit_bar,
            'timestamp': datetime.now(timezone.utc), **kwargs})
        if len(self.trade_history) > 200:
            self.trade_history = self.trade_history[-200:]

    def get_performance_stats(self):
        n = self.total_trades
        if n == 0: return {'total_trades': 0, 'win_rate': 0, 'total_profit': 0}
        wr = self.winning_trades / n * 100
        wins = [t['profit'] for t in self.trade_history if t['profit'] > 0]
        losses = [t['profit'] for t in self.trade_history if t['profit'] < 0]
        return {'total_trades': n, 'win_rate': wr, 'total_profit': self.total_profit,
                'avg_profit': self.total_profit / n,
                'avg_win': float(np.mean(wins)) if wins else 0,
                'avg_loss': float(abs(np.mean(losses))) if losses else 0,
                'winning_trades': self.winning_trades, 'losing_trades': self.losing_trades}

    def execute_buy(self, shares, price, atr=1., quality_score=70, tier=1, **kwargs):
        if self._pending_signal is not None:
            direction = self._pending_signal.get('direction', 'long')
        else:
            direction = 'long'

        use_maker = getattr(self, 'use_maker_orders', True)
        maker_offset = getattr(self, 'maker_order_offset_pct', 0.001)
        timeout_bars = getattr(self, 'maker_order_timeout_bars', 3)

        if use_maker:
            if direction == 'long':
                limit_price = price * (1 - maker_offset)
                order_type = 'buy_limit'
            else:
                limit_price = price * (1 + maker_offset)
                order_type = 'sell_limit'

            self._log(f"📊 MAKER ORDER: {direction.upper()} @ ${limit_price:.4f} "
                      f"(market ${price:.4f}, offset {maker_offset*100:.2f}%)", "cyan")

            success = self.trading_app.place_order(
                order_type,
                limit_price,
                quantity=shares,
                atr=atr,
                confidence=quality_score,
                exit_reason=None,
                position_intent='open_long' if direction == 'long' else 'open_short',
                limit_price=limit_price,
                time_in_force='GTC'
            )

            if success:
                self._maker_order_pending = True
                self._maker_order_timeout = self.bar_count + timeout_bars
                self._maker_order_price = limit_price
                self._maker_order_direction = direction
        else:
            success = self.trading_app.place_order(
                'buy' if direction == 'long' else 'sell',
                price,
                quantity=shares,
                atr=atr,
                confidence=quality_score,
                exit_reason=None,
                position_intent='open_long' if direction == 'long' else 'open_short'
            )

        if success:
            atr_mult = self._tier_stop_mult(tier, direction)
            stop = (price - atr * atr_mult if direction == 'long'
                    else price + atr * atr_mult)

            self.position = {
                'type': direction,
                'entry_price': price,
                'quantity': shares,
                'original_quantity': shares,
                'stop_loss': stop,
                'trailing_stop': stop,
                'trailing_activated': False,
                'highest_price': price if direction == 'long' else None,
                'lowest_price': price if direction == 'short' else None,
                'entry_bar': self.bar_count,
                'partial_exits': 0,
                'tier': tier,
                'entry_time': datetime.now(timezone.utc),
                'entry_quality_score': quality_score,
                'entry_reason': 'signal',
                'trade_id': self.trade_counter + 1,
                'partial_pnl_realised': 0.,
            }
            self.trade_counter += 1
            self.bars_held = 0
            self._transition_to_in_trade()
            self._log(
                f"✅ {'long' if direction == 'long' else 'short'} OPENED "
                f"T{tier} Q={quality_score} @${price:.4f} "
                f"SL=${stop:.4f} sz={shares:.4f}",
                "green" if direction == 'long' else "red"
            )
            return True, shares, self.trade_counter
        return False, 0, None

    def execute_sell(self, reason="manual", exit_percentage=1.0, **kwargs):
        if self.strategy_state != StrategyState.IN_TRADE:
            self._log("⚠️ execute_sell: no open position", "orange")
            return False, 0., 0.

        current_price = self.trading_app.get_current_price()
        if current_price is None and self.trading_app.current_data is not None:
            current_price = float(self.trading_app.current_data.get('Close', 0))
        if not current_price:
            self._log("❌ execute_sell: cannot determine current price", "red")
            return False, 0., 0.

        qty = self.position.get('quantity', 0)
        close_qty = qty * exit_percentage
        ep = self.position.get('entry_price', current_price)
        ptype = self.position.get('type', 'long')

        use_maker = getattr(self, 'use_maker_orders', True)
        maker_offset = getattr(self, 'maker_order_offset_pct', 0.001)

        if use_maker and exit_percentage >= 1.0:
            if ptype == 'long':
                limit_price = current_price * (1 + maker_offset)
                side = 'sell_limit'
            else:
                limit_price = current_price * (1 - maker_offset)
                side = 'buy_limit'

            success = self.trading_app.place_order(
                side,
                limit_price,
                quantity=close_qty,
                exit_reason=reason,
                limit_price=limit_price,
                time_in_force='GTC'
            )
        else:
            side = 'sell' if ptype == 'long' else 'buy'
            success = self.trading_app.place_order(
                side, current_price, quantity=close_qty, exit_reason=reason)

        if success:
            profit = ((current_price - ep) * close_qty if ptype == 'long'
                      else (ep - current_price) * close_qty)
            profit_pct = profit / (ep * close_qty) * 100 if ep and close_qty else 0.

            if exit_percentage >= 1.0:
                self.record_trade(
                    profit=profit + self.position.get('partial_pnl_realised', 0.),
                    exit_reason=reason, tier=self.position.get('tier'),
                    size=qty, direction=ptype,
                    entry_quality=self.position.get('entry_quality_score'),
                    entry_price=ep, exit_price=current_price,
                    hold_duration=self.bars_held,
                    entry_bar=self.position.get('entry_bar'),
                    exit_bar=self.bar_count)
                self._log(
                    f"{'✅' if profit > 0 else '❌'} "
                    f"{'long' if ptype == 'long' else 'short'} CLOSED "
                    f"@${current_price:.4f} P&L ${profit:+.2f} "
                    f"({profit_pct:+.2f}%) — {reason}",
                    "green" if profit > 0 else "red")
                self.position = {
                    'type': None, 'entry_price': None, 'quantity': None, 'stop_loss': None,
                    'trailing_stop': None, 'trailing_activated': False,
                    'highest_price': None, 'lowest_price': None, 'entry_bar': None,
                    'partial_exits': 0, 'original_quantity': None, 'tier': None,
                    'entry_time': None, 'entry_quality_score': None,
                    'entry_reason': None, 'trade_id': None, 'partial_pnl_realised': 0.,
                }
                self.bars_held = 0
                self._transition_to_seeking_entry()
                self.trading_app.update_status_indicators("parking")
            else:
                remaining = qty - close_qty
                self.position['quantity'] = remaining
                self.position['partial_exits'] = self.position.get('partial_exits', 0) + 1
                self.position['partial_pnl_realised'] = (
                        self.position.get('partial_pnl_realised', 0.) + profit)
                self._log(
                    f"📊 PARTIAL EXIT {exit_percentage:.0%} @${current_price:.4f} "
                    f"P&L ${profit:+.2f} | remaining {remaining:.4f}",
                    "cyan")

            return True, profit, current_price
        return False, 0., current_price

    def get_strategy_info(self) -> dict:
        t1l = getattr(self, 'quality_tier1_min_long', 80)
        t1s = getattr(self, 'quality_tier1_min_short', 80)
        t2l = getattr(self, 'quality_tier2_min_long', 72)
        t2s = getattr(self, 'quality_tier2_min_short', 72)
        return {
            'name': f"Professional Scalping v1.7.3",
            'version': "1.7.3",
            'tier_system': f"Tier 1 (Q≥{t1l}L/{t1s}S) / Tier 2 (Q≥{t2l}L/{t2s}S)",
            'expected_trades_monthly': "30-50",
            'target_win_rate': "47-55%",
            'target_cagr': "35-60%",
            'target_sharpe': "1.2-2.0",
            'max_drawdown': "<10%",
            'tier1_description': f"High-conviction Q≥{t1l}(L)/{t1s}(S), full-stack + all "
                                 f"filters, full size, tightest stop "
                                 f"({getattr(self, 'tier1_stop_mult_long', 1.8)}x/"
                                 f"{getattr(self, 'tier1_stop_mult_short', 2.4)}x ATR L/S), "
                                 f"risk {getattr(self, 'tier1_risk_pct', 0.010):.1%}/trade",
            'tier2_description': f"Standard Q≥{t2l}(L)/{t2s}(S), reduced size "
                                 f"({getattr(self, 'tier2_size_mult', 0.65):.0%}), wider stop "
                                 f"({getattr(self, 'tier2_stop_mult_long', 2.2)}x/"
                                 f"{getattr(self, 'tier2_stop_mult_short', 2.8)}x ATR L/S), "
                                 f"risk {getattr(self, 'tier2_risk_pct', 0.006):.1%}/trade",
            'timeframe': self.timeframe,
            'trade_direction': self.trade_direction,
            'maker_orders': "Enabled" if getattr(self, 'use_maker_orders', True) else "Disabled",
            'full_exit_only': "Enabled" if getattr(self, 'full_exit_at_r2', True) else "Disabled",
            'smart_partials': "Enabled" if getattr(self, 'allow_partial_exits', False) else "Disabled",
        }


# ═══════════════════════════════════════════════════════════════════════════
# GUI DIRECTION INTEGRATION FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def setup_gui_direction(gui_app):
    if hasattr(gui_app, 'trade_direction_var'):
        gui_dir = gui_app.trade_direction_var.get()
        DIRECTION_MANAGER.initialize(gui_dir)
    else:
        DIRECTION_MANAGER.initialize("both")

    if hasattr(gui_app, 'trade_direction_var'):
        def on_direction_change(*args):
            new_dir = gui_app.trade_direction_var.get()
            DIRECTION_MANAGER.set_direction(new_dir)
            print(f"🔄 GUI DIRECTION CHANGED: {DIRECTION_MANAGER.get_direction().upper()}")

            if hasattr(gui_app, 'strategy') and gui_app.strategy:
                if hasattr(gui_app.strategy, 'update_trade_direction'):
                    gui_app.strategy.update_trade_direction(new_dir)

        old_cbname = getattr(gui_app, '_direction_trace_cbname', None)
        if old_cbname:
            try:
                gui_app.trade_direction_var.trace_remove('write', old_cbname)
            except Exception:
                pass

        gui_app._direction_trace_cbname = gui_app.trade_direction_var.trace_add('write', on_direction_change)

    print(f"✅ GUI DIRECTION INTEGRATION: {DIRECTION_MANAGER.get_direction().upper()}")
    return DIRECTION_MANAGER.get_direction()


def get_gui_direction():
    return DIRECTION_MANAGER.get_direction()


def update_gui_direction(new_direction: str):
    DIRECTION_MANAGER.set_direction(new_direction)
    print(f"📊 DIRECTION UPDATED: {new_direction.upper()}")
    return new_direction


# ═══════════════════════════════════════════════════════════════════════════
# LIVE STRATEGY — v1.7.3
# ═══════════════════════════════════════════════════════════════════════════

class ScalpingStrategy(BaseStrategy, ScalpingLogic):
    """
    v1.7.3 Live Scalping — tiered (T1/T2) x direction risk engine.
    Commission-optimized: smart partial exits, higher quality thresholds, maker orders.
    Direction is ALWAYS taken from GUI via DirectionManager.
    """

    _instance_active: bool = False

    @classmethod
    def _check_and_claim_instance(cls) -> None:
        if cls._instance_active:
            raise RuntimeError(
                "ScalpingStrategy: another instance is already running. "
                "Stop the current session before starting a new one.")
        cls._instance_active = True

    @classmethod
    def _release_instance(cls) -> None:
        cls._instance_active = False

    def __init__(self, trading_app=None):
        self.__class__._check_and_claim_instance()
        try:
            BaseStrategy.__init__(self, trading_app)
            self.timeframe = (trading_app.get_timeframe()
                              if trading_app and hasattr(trading_app, 'get_timeframe')
                              else GlobalConfig.ACTIVE_TIMEFRAME)
            GlobalConfig.update_timeframe(self.timeframe)
            override = (trading_app.get_current_momentum_params()
                        if trading_app and hasattr(trading_app, 'get_current_momentum_params')
                        else None)
            params = ScalpingConfig.get_config(override, self.timeframe)

            if trading_app and hasattr(trading_app, 'trade_direction_var'):
                gui_dir = _normalize_trade_direction(trading_app.trade_direction_var.get())
            else:
                gui_dir = _normalize_trade_direction(params.get('trade_direction', 'both'))
            params['trade_direction'] = gui_dir
            DIRECTION_MANAGER.set_direction(gui_dir)

            ScalpingLogic.__init__(self, config=params, trading_app=trading_app)

            if trading_app and hasattr(trading_app, 'trade_direction_var'):
                gui_dir = _normalize_trade_direction(trading_app.trade_direction_var.get())
                self.trade_direction = gui_dir
                self.only_long_entries = (gui_dir == 'long')
                self.only_short_entries = (gui_dir == 'short')
                self.config['trade_direction'] = gui_dir
                DIRECTION_MANAGER.set_direction(gui_dir)
                if hasattr(self, '_last_dir_cache'):
                    self._last_dir_cache = {'bar': -999, 'dir': None, 'result': True, 'reason': ''}
                self._log(f"🔧 DIRECTION SET: {gui_dir.upper()}", "bold green")

            self.name = f"Professional Scalping v1.7.3 — {self.timeframe}"
            self.position = {
                'type': None, 'entry_price': None, 'quantity': None,
                'stop_loss': None, 'trailing_stop': None, 'trailing_activated': False,
                'highest_price': None, 'lowest_price': None, 'entry_bar': None,
                'partial_exits': 0, 'original_quantity': None, 'tier': None,
                'entry_time': None, 'entry_quality_score': None,
                'entry_reason': None, 'trade_id': None, 'partial_pnl_realised': 0.,
            }
            self.bars_held = 0
            self.trade_counter = 0
            if self.trading_app:
                self._log("=" * 70, "cyan")
                self._log(f"SCALPING v1.7.3 — {self.timeframe.upper()} LIVE", "bold green")
                self._log(f"  ✅ Direction Filter: {self.trade_direction.upper()} (from GUI)", "bold cyan")
                self._log(f"  ✅ DirectionManager: {DIRECTION_MANAGER.get_direction().upper()}", "bold cyan")
                self._log(f"  ✅ Stop T1 {params['tier1_stop_mult_long']}x/{params['tier1_stop_mult_short']}x "
                          f"| T2 {params['tier2_stop_mult_long']}x/{params['tier2_stop_mult_short']}x ATR (L/S)",
                          "green")
                self._log(f"  ✅ T1 Q≥{params['quality_tier1_min_long']}(L)/≥{params['quality_tier1_min_short']}(S) "
                          f"| T2 Q≥{params['quality_tier2_min_long']}(L)/≥{params['quality_tier2_min_short']}(S) "
                          f"| Age≥{params['trend_age_min_bars']}b "
                          f"| PZ {params['pullback_zone_lower_pct']}/+{params['pullback_zone_upper_pct']}%", "green")
                self._log(f"  ✅ ADX≥{params['adx_min_long']} | Vol≥{params['volume_min_ratio']}x "
                          f"| MaxTrades/day:{params['max_daily_trades']}", "green")
                self._log(f"  ✅ SPIKE regime blocking active (ATR >{params['atr_spike_ratio']}×avg)", "green")
                self._log(f"  ✅ SMART PARTIAL EXITS: {'ON' if params.get('allow_partial_exits', False) else 'OFF'}", "green")
                self._log(f"  ✅ MAKER ORDERS: {'ON' if params.get('use_maker_orders', True) else 'OFF'}", "green")
                self._log(f"  ✅ Single-instance guard: active", "green")
                self._log("=" * 70, "cyan")
        except Exception:
            self.__class__._release_instance()
            raise

    def __del__(self):
        self.__class__._release_instance()

    def on_stop(self):
        self.__class__._release_instance()
        if hasattr(super(), 'on_stop'):
            super().on_stop()

    def run_analysis_cycle(self, current_data, current_price, df=None):
        self._current_df = df

        if getattr(self, '_maker_order_pending', False):
            if self.bar_count >= getattr(self, '_maker_order_timeout', self.bar_count + 3):
                self._log(f"⏰ MAKER ORDER TIMEOUT: {getattr(self, '_maker_order_direction', 'unknown').upper()} "
                          f"@ ${getattr(self, '_maker_order_price', 0):.4f} cancelled", "orange")
                self._maker_order_pending = False
                if hasattr(self.trading_app, 'cancel_order'):
                    self.trading_app.cancel_order()

        if self.strategy_state == StrategyState.SEEKING_ENTRY:
            return self.check_entry_conditions(current_data)
        return self.check_exit_conditions(current_data, current_price)

    def calculate_indicators(self, df):
        return IndicatorCalculator.calculate(df, self.config)

    def check_entry_conditions(self, current_data):
        if self.strategy_state != StrategyState.SEEKING_ENTRY:
            return "hold", 0, 0, "in_trade"
        action, quality, reason, components = self._check_entry_conditions(current_data)
        if self._pending_signal is not None and self.bar_count > self._signal_bar:
            sig = self._pending_signal
            ep = current_data.get('Close', 0)
            size = self.calculate_position_size(
                self.risk_controller.current_equity
                if hasattr(self, 'risk_controller') else GlobalConfig.INITIAL_CAPITAL,
                current_data.get('ATR', 1.), ep,
                sig['quality_score'], sig['tier'], sig.get('position_mult', 1.))
            if size > 0:
                d = sig['direction']
                self._pending_signal = None
                self._signal_bar = -999
                qs = sig['quality_score']
                if d == 'long':
                    return ("buy", qs, size, f"pQ{qs}")
                else:
                    return ("sell_short", qs, size, f"pQ{qs}")
        return action, quality, 0, reason

    def check_exit_conditions(self, current_data, current_price):
        if self.strategy_state != StrategyState.IN_TRADE: return None, 1.
        if self.position['type'] == 'long':
            self.position['highest_price'] = max(
                self.position.get('highest_price') or current_price, current_price)
        else:
            lp = self.position.get('lowest_price')
            self.position['lowest_price'] = min(lp, current_price) if lp is not None else current_price

        atr = current_data.get('ATR')
        if not atr or atr <= 0:
            raise ValueError(f"Invalid ATR bar {self.bars_held}")

        tier_activation, tier_distance = self._tier_trailing_params(self.position.get('tier') or 2)
        ta = max((atr / current_price) * 1.5, tier_activation)
        td = max((atr / current_price) * 1.0, tier_distance)
        if self.position['type'] == 'long':
            pp = (current_price - self.position['entry_price']) / self.position['entry_price']
            if not self.position['trailing_activated'] and pp >= ta:
                self.position['trailing_activated'] = True
                self.position['trailing_stop'] = current_price * (1 - td)
            if self.position['trailing_activated']:
                ns = self.position['highest_price'] * (1 - td)
                if ns > (self.position['trailing_stop'] or 0):
                    self.position['trailing_stop'] = ns
        else:
            pp = (self.position['entry_price'] - current_price) / self.position['entry_price']
            if not self.position['trailing_activated'] and pp >= ta:
                self.position['trailing_activated'] = True
                self.position['trailing_stop'] = current_price * (1 + td)
            if self.position['trailing_activated']:
                ns = self.position['lowest_price'] * (1 + td)
                if ns < (self.position['trailing_stop'] or float('inf')):
                    self.position['trailing_stop'] = ns

        exit_signal, exit_pct = self.exit_manager.evaluate_exit(
            current_price=current_price,
            entry_price=self.position['entry_price'],
            stop_loss=self.position['stop_loss'],
            highest_price=self.position.get('highest_price', current_price),
            lowest_price=self.position.get('lowest_price', current_price),
            bars_held=self.bars_held,
            partial_exits=self.position.get('partial_exits', 0),
            ema_fast=current_data.get('EMA_Fast', 0),
            ema_mid=current_data.get('EMA_Mid', 0),
            ema_slow=current_data.get('EMA_Slow', 0),
            macd=current_data.get('MACD', 0),
            macd_signal=current_data.get('MACD_Signal', 0),
            macd_prev=current_data.get('MACD_closed', 0),
            signal_prev=current_data.get('MACD_Signal_closed', 0),
            stoch_k=current_data.get('Stoch_K', 50),
            stoch_d=current_data.get('Stoch_D', 50),
            rsi=current_data.get('RSI', 50),
            adx=current_data.get('ADX', 0),
            atr=atr,
            position_type=self.position['type'],
            trailing_activated=self.position.get('trailing_activated', False),
            trailing_stop=self.position.get('trailing_stop'),
            tier=self.position.get('tier') or 2,
            original_quantity=self.position.get('original_quantity', 100),
            remaining_quantity=self.position.get('quantity', 100)
        )

        return exit_signal, exit_pct

    def on_bar_update(self, current_equity):
        self.equity_curve.append(current_equity)
        self.bar_count += 1

        if getattr(self, '_maker_order_pending', False):
            if self.bar_count >= getattr(self, '_maker_order_timeout', self.bar_count + 3):
                self._log(f"⏰ MAKER ORDER TIMEOUT: {getattr(self, '_maker_order_direction', 'unknown').upper()} "
                          f"@ ${getattr(self, '_maker_order_price', 0):.4f} cancelled", "orange")
                self._maker_order_pending = False
                if hasattr(self.trading_app, 'cancel_order'):
                    self.trading_app.cancel_order()

        if self.strategy_state == StrategyState.IN_TRADE:
            self.bars_held += 1

        self.risk_controller.current_equity = current_equity
        if current_equity > self.risk_controller.peak_equity:
            self.risk_controller.peak_equity = current_equity

    def get_strategy_stats(self):
        s = self.get_performance_stats()
        r = self.risk_controller.get_stats()
        return {'strategy_name': self.name, 'timeframe': self.timeframe,
                'total_trades': s['total_trades'], 'win_rate': s.get('win_rate', 0),
                'total_profit': s.get('total_profit', 0), 'profit_factor': r['profit_factor'],
                'max_drawdown': r['max_drawdown'], 'sharpe_ratio': r['sharpe_ratio'],
                'current_regime': self.current_regime,
                'strategy_state': self.strategy_state.name,
                'trade_direction': self.trade_direction}


# ═══════════════════════════════════════════════════════════════════════════
# BACKTEST STRATEGY — v1.7.3
# ═══════════════════════════════════════════════════════════════════════════

class BacktestScalpingStrategy(Strategy, ScalpingLogic):
    """
    v1.7.3 Backtest — tiered (T1/T2) x direction risk engine.
    Commission-optimized: smart partial exits, higher quality thresholds.
    Direction is ALWAYS taken from GUI via DirectionManager.
    """

    for _k in SCALPING_PARAMS:
        locals()[_k] = None
    del _k

    _use_updated_params = False
    _updated_params = {}
    _timeframe = "1h"

    @classmethod
    def set_updated_params(cls, params=None, timeframe=None, trade_direction=None):
        if params and isinstance(params, dict):
            cls._use_updated_params = True
            cls._updated_params = params.copy()
        elif not cls._updated_params:
            cls._updated_params = {}

        if trade_direction:
            cls._updated_params['trade_direction'] = _normalize_trade_direction(trade_direction)

        if 'trade_direction' in cls._updated_params:
            cls._updated_params['trade_direction'] = _normalize_trade_direction(
                cls._updated_params['trade_direction'])
            DIRECTION_MANAGER.set_direction(cls._updated_params['trade_direction'])

        if timeframe:
            cls._timeframe = timeframe
            GlobalConfig.update_timeframe(timeframe)

        print(f"\n{'=' * 60}")
        print(f"✅ SCALPING BACKTEST DIRECTION: {cls._updated_params.get('trade_direction', 'both').upper()}")
        print(f"{'=' * 60}\n")

    @classmethod
    def reset_to_defaults(cls):
        cls._use_updated_params = False
        cls._updated_params = {}
        cls._timeframe = "1h"

    def __init__(self, broker, data, params):
        Strategy.__init__(self, broker, data, params)
        try:
            self.timeframe = (data.timeframe
                              if hasattr(data, 'timeframe')
                              else params.get('timeframe', self.__class__._timeframe))
        except:
            self.timeframe = self.__class__._timeframe

        GlobalConfig.update_timeframe(self.timeframe)
        override = self.__class__._updated_params if self.__class__._use_updated_params else {}
        config = ScalpingConfig.get_config(override, self.timeframe)

        if params:
            for k, v in params.items():
                if k in config:
                    config[k] = v

        self._trade_direction = _normalize_trade_direction(config.get('trade_direction', 'both'))
        config['trade_direction'] = self._trade_direction
        self.only_long_entries = (self._trade_direction == 'long')
        self.only_short_entries = (self._trade_direction == 'short')
        self.trade_direction = self._trade_direction
        DIRECTION_MANAGER.set_direction(self._trade_direction)

        for k, v in config.items():
            setattr(self, k, v)
        ScalpingLogic.__init__(self, config=config, trading_app=None)

        print(f"\n{'=' * 68}")
        print(f"  BACKTEST SCALPING v1.7.3 — {self.timeframe.upper()} (SMART PARTIAL EXITS)")
        print(f"{'=' * 68}")
        print(f"  Direction  : {self._trade_direction.upper()}")
        print(f"  Quality T1 : ≥{config['quality_tier1_min_long']}(L) / ≥{config['quality_tier1_min_short']}(S)")
        print(f"  Quality T2 : ≥{config['quality_tier2_min_long']}(L) / ≥{config['quality_tier2_min_short']}(S)")
        print(f"  Stop T1    : {config['tier1_stop_mult_long']}x(L)/{config['tier1_stop_mult_short']}x(S) ATR")
        print(f"  Stop T2    : {config['tier2_stop_mult_long']}x(L)/{config['tier2_stop_mult_short']}x(S) ATR")
        print(f"  Smart Partials: {'ON' if config.get('allow_partial_exits', False) else 'OFF'}")
        print(f"  Maker Orders: {'ON' if config.get('use_maker_orders', True) else 'OFF'}")
        print(f"{'=' * 68}\n")

        self._entry_price = np.nan
        self._stop_loss = np.nan
        self._highest_price = np.nan
        self._lowest_price = np.nan
        self._trailing_activated = False
        self._trailing_stop = None
        self._be_stop_set = False
        self._bars_held = 0
        self._partial_exits = 0
        self._entry_bar = -999
        self._entry_tier = None
        self._entry_quality = 0
        self._partial_pnl_realised = 0.
        self._position_direction = 'long'
        self._pending_signal = None
        self._signal_bar = -999
        self._exit_reason_map = {}

        self._maker_order_pending = False
        self._maker_order_timeout = -999
        self._maker_order_price = 0.0
        self._maker_order_direction = ''
        self._maker_order_pending_bar = -999

        try:
            delta = int((data.df.index[1] - data.df.index[0]).total_seconds() / 60)
        except:
            delta = int(GlobalConfig.TIMEFRAME_MINUTES.get(self.timeframe, 60))
        self.bar_interval_minutes = delta
        self._bar_interval_minutes = delta
        self.config['bar_interval_minutes'] = delta

    def _should_trade_direction(self, direction: str) -> bool:
        current_direction = self._trade_direction

        if current_direction == 'both':
            return True
        if current_direction == 'long':
            return direction == 'long'
        if current_direction == 'short':
            return direction == 'short'

        return True

    def _check_entry_conditions(self, data):
        my_dir = self._trade_direction

        if my_dir == 'long':
            return self._entry_long(data)
        elif my_dir == 'short':
            return self._entry_short(data)
        else:
            return self._entry_both(data)

    def _entry_both(self, data):
        my_dir = self._trade_direction

        if my_dir == 'long':
            return self._entry_long(data)
        elif my_dir == 'short':
            return self._entry_short(data)

        sp = self._pending_signal
        sb_ = self._signal_bar
        lr = self._entry_long(data)
        lp = self._pending_signal
        lb_ = self._signal_bar
        self._pending_signal = sp
        self._signal_bar = sb_
        sr = self._entry_short(data)
        short_p = self._pending_signal
        short_b = self._signal_bar

        la, lq, ll, lc = lr
        sa, sq, sl_, sc = sr

        lpp = bool(ll and ll.startswith("PENDING:"))
        spp = bool(sl_ and sl_.startswith("PENDING:"))
        lq = lq if isinstance(lq, (int, float)) else 0
        sq = sq if isinstance(sq, (int, float)) else 0

        if lpp and spp:
            if lq >= sq:
                self._pending_signal = lp
                self._signal_bar = lb_
                self._log(f"⚖️ long Q{lq}>Q{sq}", "cyan")
                return lr
            else:
                self._pending_signal = short_p
                self._signal_bar = short_b
                self._log(f"⚖️ short Q{sq}>Q{lq}", "cyan")
                return sr

        if lpp:
            self._pending_signal = lp
            self._signal_bar = lb_
            return lr
        if spp:
            self._pending_signal = short_p
            self._signal_bar = short_b
            return sr

        self._pending_signal = sp
        self._signal_bar = sb_

        if lq >= sq and lq > 0:
            reason = ll if ll else "no_clear_dir"
            return "hold", lq, reason, lc
        elif sq > 0:
            reason = sl_ if sl_ else "no_clear_dir"
            return "hold", sq, reason, sc
        return "hold", 0, "no_clear_dir", {}

    def _bt_safe_size(self, units, price):
        if units <= 0:
            return 0
        if units >= 1 and units == round(units):
            return int(units)
        return max(.0001, min((units * price) / max(self.equity, 1.), .9999))

    def export_trades_to_excel(self, filename: str) -> bool:
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill, Font

            if not self.trade_history:
                print("⚠️ export_trades_to_excel: no trades to export")
                return False

            rows = []
            for i, t in enumerate(self.trade_history, start=1):
                profit = t.get('profit', 0.)
                ep = t.get('entry_price', 0.)
                xp = t.get('exit_price', 0.)
                sz = t.get('size', 0.)
                direction = t.get('direction', 'long')
                entry_val = ep * sz if ep and sz else 0.
                ret_pct = profit / entry_val * 100 if entry_val else 0.
                hold = t.get('hold_duration', 0)
                ts = t.get('timestamp', '')
                if hasattr(ts, 'strftime'):
                    ts = ts.strftime('%Y-%m-%d %H:%M:%S')
                rows.append({
                    'Trade_#': i,
                    'Direction': direction.upper(),
                    'Entry_Bar': t.get('entry_bar', ''),
                    'Exit_Bar': t.get('exit_bar', ''),
                    'Timestamp': ts,
                    'Entry_Price': round(ep, 4),
                    'Exit_Price': round(xp, 4),
                    'Size': round(sz, 4),
                    'Profit_$': round(profit, 2),
                    'Return_%': round(ret_pct, 2),
                    'Hold_Bars': hold,
                    'Exit_Reason': t.get('exit_reason', ''),
                    'Tier': t.get('tier', ''),
                    'Quality_Score': t.get('entry_quality', ''),
                    'Win': 'Yes' if profit > 0 else 'No',
                })

            df_trades = pd.DataFrame(rows)

            n = len(df_trades)
            wins = (df_trades['Profit_$'] > 0).sum()
            losses = n - wins
            gp = df_trades.loc[df_trades['Profit_$'] > 0, 'Profit_$'].sum()
            gl = abs(df_trades.loc[df_trades['Profit_$'] < 0, 'Profit_$'].sum())
            pf = round(gp / gl, 3) if gl > 0 else float('inf')

            long_trades = df_trades[df_trades['Direction'] == 'long']
            short_trades = df_trades[df_trades['Direction'] == 'short']

            tier1_trades = df_trades[df_trades['Tier'] == 1]
            tier2_trades = df_trades[df_trades['Tier'] == 2]

            summary_rows = [
                ['Metric', 'Value'],
                ['Total Trades', n],
                ['Win Rate', f"{wins / n * 100:.1f}%" if n else '0%'],
                ['Profit Factor', pf if pf != float('inf') else '∞'],
                ['Total Profit $', round(df_trades['Profit_$'].sum(), 2)],
                ['Avg Profit / Trade $', round(df_trades['Profit_$'].mean(), 2)],
                ['Best Trade $', round(df_trades['Profit_$'].max(), 2)],
                ['Worst Trade $', round(df_trades['Profit_$'].min(), 2)],
                ['long trades', len(long_trades)],
                ['short trades', len(short_trades)],
                ['long win rate', f"{(long_trades['Profit_$'] > 0).mean() * 100:.1f}%"
                if len(long_trades) else 'N/A'],
                ['short win rate', f"{(short_trades['Profit_$'] > 0).mean() * 100:.1f}%"
                if len(short_trades) else 'N/A'],
                ['Tier 1 trades', len(tier1_trades)],
                ['Tier 1 win rate', f"{(tier1_trades['Profit_$'] > 0).mean() * 100:.1f}%"
                if len(tier1_trades) else 'N/A'],
                ['Tier 2 trades', len(tier2_trades)],
                ['Tier 2 win rate', f"{(tier2_trades['Profit_$'] > 0).mean() * 100:.1f}%"
                if len(tier2_trades) else 'N/A'],
                ['Timeframe', getattr(self, 'timeframe', '1h')],
                ['Direction', getattr(self, '_trade_direction', 'both')],
                ['Tier1 Stop Mult (L/S)', f"{getattr(self, 'tier1_stop_mult_long', 1.8)}x / "
                                          f"{getattr(self, 'tier1_stop_mult_short', 2.4)}x"],
                ['Tier2 Stop Mult (L/S)', f"{getattr(self, 'tier2_stop_mult_long', 2.2)}x / "
                                          f"{getattr(self, 'tier2_stop_mult_short', 2.8)}x"],
                ['Tier1 Risk % / Trade', f"{getattr(self, 'tier1_risk_pct', 0.010):.2%}"],
                ['Tier2 Risk % / Trade', f"{getattr(self, 'tier2_risk_pct', 0.006):.2%}"],
                ['Quality Tier1 Min (L/S)', f"{getattr(self, 'quality_tier1_min_long', 80)} / "
                                            f"{getattr(self, 'quality_tier1_min_short', 80)}"],
                ['Quality Tier2 Min (L/S)', f"{getattr(self, 'quality_tier2_min_long', 72)} / "
                                            f"{getattr(self, 'quality_tier2_min_short', 72)}"],
                ['Smart Partials', 'ON' if getattr(self, 'allow_partial_exits', False) else 'OFF'],
                ['Maker Orders', 'ON' if getattr(self, 'use_maker_orders', True) else 'OFF'],
            ]
            df_summary = pd.DataFrame(summary_rows[1:], columns=summary_rows[0])

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_trades.to_excel(writer, sheet_name='Trades', index=False)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)

                wb = writer.book
                ws = wb['Trades']
                green_fill = PatternFill("solid", fgColor="C6EFCE")
                red_fill = PatternFill("solid", fgColor="FFC7CE")
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    profit_cell = next(
                        (c for c in row if ws.cell(1, c.column).value == 'Profit_$'), None)
                    if profit_cell is None:
                        continue
                    fill = (green_fill if (profit_cell.value or 0) > 0
                            else red_fill)
                    for cell in row:
                        cell.fill = fill

            print(f"✅ Scalping trades exported: {filename} ({n} trades)")
            return True

        except Exception as e:
            import traceback
            print(f"❌ export_trades_to_excel error: {e}")
            print(traceback.format_exc())
            return False

    def init(self):
        df = IndicatorCalculator.calculate(self.data.df.copy(), self.config)
        self.df_indicators = df
        self.df_enhanced = df.copy()

    def next(self):
        idx = len(self.data) - 1
        cd = self._build_current_data()
        cp = float(self.data.Close[-1])

        if cd['ATR'] <= 0 or cd['ADX'] == 0:
            return

        self._current_df = self.df_enhanced.iloc[:idx + 1]
        self.bar_count = idx

        if getattr(self, '_maker_order_pending', False):
            if self.bar_count >= getattr(self, '_maker_order_timeout', self.bar_count + 3):
                self._maker_order_pending = False
                self._maker_order_pending_bar = -999
                print(f"⏰ [BACKTEST] MAKER ORDER TIMEOUT: {getattr(self, '_maker_order_direction', 'unknown').upper()} "
                      f"@ ${getattr(self, '_maker_order_price', 0):.4f}")

        if self._pending_signal is not None and self.bar_count > self._signal_bar:
            sig = self._pending_signal
            direction = sig.get('direction', 'long')

            if not self._should_trade_direction(direction):
                print(
                    f"⏸️ DIRECTION FILTER BLOCKED: {direction.upper()} entry (only {self._trade_direction.upper()} allowed)")
                self._pending_signal = None
                self._signal_bar = -999
                return

            ep = float(self.data.Open[-1]) if not np.isnan(self.data.Open[-1]) else cp
            self._position_direction = 'long' if sig['decision'] == 'buy' else 'short'

            size = self.calculate_position_size(
                self.equity,
                cd['ATR'],
                ep,
                sig['quality_score'],
                sig['tier'],
                sig.get('position_mult', 1.),
                direction=self._position_direction
            )

            if size > 0:
                sd = self._tier_stop_mult(sig['tier'], self._position_direction) * cd['ATR']
                if self._position_direction == 'long':
                    stop = ep - sd
                    if stop < ep:
                        use_maker = getattr(self, 'use_maker_orders', True)
                        maker_offset = getattr(self, 'maker_order_offset_pct', 0.001)
                        if use_maker:
                            limit_price = ep * (1 - maker_offset)
                            if self.data.Low[-1] <= limit_price:
                                self.buy(size=self._bt_safe_size(size, limit_price), limit=limit_price)
                                self._maker_order_pending = False
                            else:
                                self._maker_order_pending = True
                                self._maker_order_timeout = self.bar_count + getattr(self, 'maker_order_timeout_bars', 3)
                                self._maker_order_price = limit_price
                                self._maker_order_direction = self._position_direction
                                self._maker_order_pending_bar = self.bar_count
                                self._pending_signal = None
                                self._signal_bar = -999
                                print(f"📊 MAKER ORDER PENDING: {self._position_direction.upper()} @ ${limit_price:.4f}")
                                return
                        else:
                            self.buy(size=self._bt_safe_size(size, ep))
                    else:
                        self._pending_signal = None
                        return
                else:
                    stop = ep + sd
                    if stop > ep:
                        use_maker = getattr(self, 'use_maker_orders', True)
                        maker_offset = getattr(self, 'maker_order_offset_pct', 0.001)
                        if use_maker:
                            limit_price = ep * (1 + maker_offset)
                            if self.data.High[-1] >= limit_price:
                                self.sell(size=self._bt_safe_size(size, limit_price), limit=limit_price)
                                self._maker_order_pending = False
                            else:
                                self._maker_order_pending = True
                                self._maker_order_timeout = self.bar_count + getattr(self, 'maker_order_timeout_bars', 3)
                                self._maker_order_price = limit_price
                                self._maker_order_direction = self._position_direction
                                self._maker_order_pending_bar = self.bar_count
                                self._pending_signal = None
                                self._signal_bar = -999
                                print(f"📊 MAKER ORDER PENDING: {self._position_direction.upper()} @ ${limit_price:.4f}")
                                return
                        else:
                            self.sell(size=self._bt_safe_size(size, ep))
                    else:
                        self._pending_signal = None
                        return

                self._entry_price = ep
                self._stop_loss = stop
                self._highest_price = ep if self._position_direction == 'long' else None
                self._lowest_price = ep if self._position_direction == 'short' else None
                self._bars_held = 0
                self._partial_exits = 0
                self._entry_bar = idx
                self._entry_tier = sig['tier']
                self._entry_quality = sig['quality_score']
                self._partial_pnl_realised = 0.
                self._trailing_activated = False
                self._trailing_stop = None
                self._be_stop_set = False
                self._transition_to_in_trade()

                icon = "⬆️" if self._position_direction == 'long' else "⬇️"
                print(f"{icon} ENTER T{sig['tier']} Q={sig['quality_score']} "
                      f"@${ep:.2f} SL=${stop:.2f} sz={size:.4f}")
                self._register_daily_trade()
            self._pending_signal = None
            self._signal_bar = -999
            return

        if self.strategy_state == StrategyState.SEEKING_ENTRY:
            self._check_entry_conditions(cd)
            return

        if self.strategy_state == StrategyState.IN_TRADE:
            self._bars_held += 1

            if self._position_direction == 'long':
                if self._highest_price is None or cp > self._highest_price:
                    self._highest_price = cp
            else:
                if self._lowest_price is None or cp < self._lowest_price:
                    self._lowest_price = cp

            atr = cd.get('ATR', 0)
            if atr <= 0:
                return

            atr_pct = atr / cp if cp > 0 else .001
            tier_activation, tier_distance = self._tier_trailing_params(self._entry_tier or 2)
            ta = max(atr_pct * 1.5, tier_activation)
            td = max(atr_pct * 1.0, tier_distance)

            if self._position_direction == 'long':
                if not self._trailing_activated:
                    if (cp - self._entry_price) / self._entry_price >= ta:
                        self._trailing_activated = True
                        self._trailing_stop = cp * (1 - td)
                if self._trailing_activated:
                    ns = self._highest_price * (1 - td)
                    if ns > (self._trailing_stop or 0):
                        self._trailing_stop = ns
            else:
                if not self._trailing_activated:
                    if (self._entry_price - cp) / self._entry_price >= ta:
                        self._trailing_activated = True
                        self._trailing_stop = cp * (1 + td)
                if self._trailing_activated:
                    ns = self._lowest_price * (1 + td)
                    if ns < (self._trailing_stop or 1e9):
                        self._trailing_stop = ns

            if getattr(self, 'be_stop_enabled', True) and not self._be_stop_set:
                sdd = abs(self._entry_price - self._stop_loss)
                if sdd > 0:
                    ber = getattr(self, 'be_stop_r_trigger', 1.5)
                    beb = getattr(self, 'be_stop_no_progress_bars', 10)
                    if self._position_direction == 'long':
                        pa = cp - self._entry_price
                        if pa >= ber * sdd:
                            self._stop_loss = self._entry_price
                            self._be_stop_set = True
                        elif self._bars_held >= beb and pa / self._entry_price < .002:
                            self._stop_loss = self._entry_price
                            self._be_stop_set = True
                    else:
                        pa = self._entry_price - cp
                        if pa >= ber * sdd:
                            self._stop_loss = self._entry_price
                            self._be_stop_set = True
                        elif self._bars_held >= beb and pa / self._entry_price < .002:
                            self._stop_loss = self._entry_price
                            self._be_stop_set = True

            es, ep2 = self.exit_manager.evaluate_exit(
                current_price=cp,
                entry_price=self._entry_price,
                stop_loss=self._stop_loss,
                highest_price=self._highest_price,
                lowest_price=self._lowest_price,
                bars_held=self._bars_held,
                partial_exits=self._partial_exits,
                ema_fast=cd['EMA_Fast'],
                ema_mid=cd['EMA_Mid'],
                ema_slow=cd['EMA_Slow'],
                macd=cd['MACD'],
                macd_signal=cd['MACD_Signal'],
                macd_prev=cd['MACD_closed'],
                signal_prev=cd['MACD_Signal_closed'],
                stoch_k=cd['Stoch_K'],
                stoch_d=cd['Stoch_D'],
                rsi=cd['RSI'],
                adx=cd['ADX'],
                atr=atr,
                position_type=self._position_direction,
                trailing_activated=self._trailing_activated,
                trailing_stop=self._trailing_stop,
                tier=self._entry_tier or 2,
                original_quantity=self.position.size + self._partial_pnl_realised / self._entry_price if self._entry_price else 0,
                remaining_quantity=abs(self.position.size) if self.position else 0
            )

            if es is None:
                mh = self.config.get('min_hold_bars_before_stop', 2)
                sd_ = atr * self._tier_stop_mult(self._entry_tier or 2, self._position_direction)
                em = self.config.get('emergency_stop_mult', 3.0)
                if self._position_direction == 'long' and cp <= self._stop_loss:
                    es = ("stop_loss_hard" if self._bars_held >= mh
                          else ("stop_loss_emergency"
                                if cp <= self._entry_price - sd_ * em else None))
                    if es:
                        ep2 = 1.
                elif self._position_direction == 'short' and cp >= self._stop_loss:
                    es = ("stop_loss_hard" if self._bars_held >= mh
                          else ("stop_loss_emergency"
                                if cp >= self._entry_price + sd_ * em else None))
                    if es:
                        ep2 = 1.

            if es is None and self._trailing_activated and self._trailing_stop is not None:
                if self._position_direction == 'long' and cp <= self._trailing_stop:
                    es = "trailing_stop"
                    ep2 = 1.
                elif self._position_direction == 'short' and cp >= self._trailing_stop:
                    es = "trailing_stop"
                    ep2 = 1.

            if es is None and idx == len(self.df_enhanced) - 1 and self._bars_held > 0:
                es = "end_of_backtest"
                ep2 = 1.
                print("🏁 FORCED EXIT")

            if es:
                pct = ((cp - self._entry_price) / self._entry_price * 100
                       if self._position_direction == 'long'
                       else (self._entry_price - cp) / self._entry_price * 100)

                self._bt_close_position(cp, es, pct)

    def _bt_close_position(self, current_price, exit_signal, profit_pct):
        sz = abs(self.position.size)
        self._exit_reason_map[self._entry_bar] = exit_signal
        self.position.close()

        fl = ((current_price - self._entry_price) * sz
              if self._position_direction == 'long'
              else (self._entry_price - current_price) * sz)
        tp = fl + self._partial_pnl_realised

        self.record_trade(
            profit=tp,
            exit_reason=exit_signal,
            tier=self._entry_tier,
            size=sz,
            direction=self._position_direction,
            entry_quality=self._entry_quality,
            entry_price=self._entry_price,
            exit_price=current_price,
            hold_duration=self._bars_held,
            entry_bar=self._entry_bar,
            exit_bar=len(self.data) - 1
        )

        icon = "⬆️" if self._position_direction == 'long' else "⬇️"
        print(f"{'✅' if tp > 0 else '❌'} {icon} EXIT@${current_price:.2f} "
              f"{profit_pct:+.2f}% hold={self._bars_held}b Q={self._entry_quality} {exit_signal}")

        self._entry_price = np.nan
        self._stop_loss = np.nan
        self._highest_price = np.nan
        self._lowest_price = np.nan
        self._bars_held = 0
        self._partial_exits = 0
        self._partial_pnl_realised = 0.
        self._trailing_activated = False
        self._trailing_stop = None
        self._be_stop_set = False
        self._entry_quality = 0
        self._transition_to_seeking_entry()

    def _build_current_data(self):
        idx = len(self.data) - 1
        row = self.df_enhanced.iloc[idx]
        cd = row.to_dict()
        cd['Open'] = float(self.data.Open[-1])
        cd['High'] = float(self.data.High[-1])
        cd['Low'] = float(self.data.Low[-1])
        cd['Close'] = float(self.data.Close[-1])
        cd['Volume'] = float(self.data.Volume[-1])
        for k in ('ATR', 'ADX', 'MACD', 'MACD_Signal', 'MACD_closed', 'MACD_Signal_closed'):
            if k not in cd or pd.isna(cd.get(k)):
                cd[k] = 0.
        return cd


# ═══════════════════════════════════════════════════════════════════════════
# GUI TIMEFRAME INTEGRATION
# ═══════════════════════════════════════════════════════════════════════════

class TimeframeManager:
    def __init__(self, parent_gui=None):
        self.parent = parent_gui
        self.current_timeframe = GlobalConfig.ACTIVE_TIMEFRAME

    def create_timeframe_selector(self, parent_frame):
        try:
            import tkinter as tk
            from tkinter import ttk
            frame = ttk.LabelFrame(parent_frame, text="Timeframe", padding=5)
            frame.pack(fill='x', padx=5, pady=5)
            self.timeframe_var = tk.StringVar(value=self.current_timeframe)
            ttk.Label(frame, text="Chart Timeframe:").pack(side='left', padx=5)
            cb = ttk.Combobox(frame, textvariable=self.timeframe_var,
                              values=GlobalConfig.VALID_TIMEFRAMES, state='readonly', width=8)
            cb.pack(side='left', padx=5)
            cb.bind('<<ComboboxSelected>>', self.on_timeframe_changed)
            return frame
        except ImportError:
            return None

    def on_timeframe_changed(self, event=None):
        self.set_timeframe(self.timeframe_var.get())

    def set_timeframe(self, tf):
        if tf not in GlobalConfig.VALID_TIMEFRAMES: return
        old = self.current_timeframe
        self.current_timeframe = tf
        GlobalConfig.update_timeframe(tf)
        print(f"📊 Timeframe: {old}→{tf} ({GlobalConfig.get_bar_interval_minutes()}min)")
        if self.parent and hasattr(self.parent, 'update_strategy_timeframe'):
            self.parent.update_strategy_timeframe(tf)
        return tf

    def get_current_timeframe(self):
        return self.current_timeframe

    def get_bar_interval_minutes(self):
        return GlobalConfig.get_bar_interval_minutes()

# ═══════════════════════════════════════════════════════════════════════════
# END OF FILE — v1.7.3
# ═══════════════════════════════════════════════════════════════════════════